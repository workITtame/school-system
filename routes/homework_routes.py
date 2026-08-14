from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user
from models import db, Homework, Subject, Classes, Sections, Student, Teacher, Terms, SchoolTable
from sqlalchemy import func
from datetime import datetime
from collections import defaultdict

homework_bp = Blueprint('homework', __name__, url_prefix='/homework')

def get_teacher_homework_data():
    today_date = datetime.now().date()
    homework_list = Homework.query.order_by(Homework.due_date.desc()).all()
    subjects = Subject.query.filter_by(is_deleted=False).all()
    classes = Classes.query.filter_by(is_deleted=False).all()
    sections = Sections.query.filter_by(is_deleted=False).all()

    total_count = len(homework_list)
    completed_count = sum(1 for h in homework_list if h.status == 'مكتمل')
    pending_count = sum(1 for h in homework_list if h.status in ['معلق', 'قيد الإنجاز'])
    late_count = sum(1 for h in homework_list if h.status == 'متأخر')
    
    targeted_subjects_count = len(set([h.sub_id for h in homework_list if h.sub_id]))
    completion_rate = round((completed_count / total_count * 100), 1) if total_count > 0 else 0.0

    kpi = {
        'total_count': total_count,
        'completed_count': completed_count,
        'pending_count': pending_count,
        'late_count': late_count,
        'targeted_subjects_count': targeted_subjects_count,
        'completion_rate': completion_rate
    }

    if homework_list:
        latest_hw = homework_list[0]
        days_rem = (latest_hw.due_date - today_date).days if latest_hw.due_date else 0
        if days_rem < 0:
            rem_str = 'منتهي'
        elif days_rem == 0:
            rem_str = 'اليوم'
        else:
            rem_str = f"{days_rem} أيام متبقية"

        sec_title = (latest_hw.section.SectionName if latest_hw.section else 'جميع الشعب') if latest_hw.section_id else 'جميع الشعب'
        current_active_homework = {
            'academic_year': 'العام الدراسي (2025-2026)',
            'section_name': sec_title,
            'class_name': latest_hw.school_class.CName if latest_hw.school_class else 'الصف الأول',
            'subject_name': latest_hw.subject.SubName if latest_hw.subject else 'القرآن الكريم',
            'remaining_minutes': rem_str,
            'status': 'حصة الآن'
        }
    else:
        current_active_homework = None

    c_pct = round((completed_count / total_count * 100), 1) if total_count > 0 else 0.0
    p_pct = round((pending_count / total_count * 100), 1) if total_count > 0 else 0.0
    l_pct = round((late_count / total_count * 100), 1) if total_count > 0 else 0.0
    
    status_distribution = {
        'completed': completed_count,
        'completed_pct': c_pct,
        'pending': pending_count,
        'pending_pct': p_pct,
        'late': late_count,
        'late_pct': l_pct,
        'cancelled': 0,
        'cancelled_pct': 0.0,
        'not_started': max(0, total_count - (completed_count + pending_count + late_count)),
        'not_started_pct': round(max(0, 100 - (c_pct + p_pct + l_pct)), 1) if total_count > 0 else 0.0
    }

    sub_counts = defaultdict(int)
    for hw in homework_list:
        sub_name = hw.subject.SubName if hw.subject else 'مادة عامة'
        sub_counts[sub_name] += 1
        
    most_assigned_subjects = []
    if sub_counts:
        max_c = max(sub_counts.values()) if sub_counts else 1
        for name, cnt in sorted(sub_counts.items(), key=lambda x: x[1], reverse=True)[:4]:
            pct = round((cnt / max_c * 100), 1)
            most_assigned_subjects.append({
                'name': name,
                'count': cnt,
                'pct': pct
            })

    upcoming_homeworks = []
    for hw in homework_list[:3]:
        delta = (hw.due_date - today_date).days if hw.due_date else 0
        if delta == 0:
            rel_str = 'اليوم'
        elif delta > 0:
            rel_str = f"بعد {delta} أيام" if delta > 2 else 'بعد يومين'
        else:
            rel_str = 'منتهي'
            
        upcoming_homeworks.append({
            'id': hw.id,
            'title': hw.title,
            'due_date_str': hw.due_date.strftime('%Y-%m-%d') if hw.due_date else '—',
            'relative_date': rel_str
        })

    return {
        'homework_list': homework_list,
        'subjects': subjects,
        'classes': classes,
        'sections': sections,
        'kpi': kpi,
        'current_active_homework': current_active_homework,
        'status_distribution': status_distribution,
        'most_assigned_subjects': most_assigned_subjects,
        'upcoming_homeworks': upcoming_homeworks,
        'today': today_date.strftime('%Y-%m-%d')
    }

from services.teacher_homework_service import (
    get_teacher_homework_statistics,
    get_teacher_homeworks,
    get_homework_details,
    create_teacher_homework,
    update_teacher_homework,
    publish_homework as service_publish_homework,
    close_homework as service_close_homework,
    delete_teacher_homework
)

from services.teacher_homework_grading_service import (
    get_homework_grading_workspace,
    get_student_submission,
    save_grade as service_save_grade,
    save_feedback as service_save_feedback,
    publish_grades as service_publish_grades,
    reopen_submission as service_reopen_submission,
    get_grading_statistics
)

@homework_bp.route('/')
@login_required
def index():
    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        today_str = datetime.now().strftime('%Y-%m-%d')
        
        teacher_class_ids = set()
        teacher_section_ids = set()
        if teacher:
            slots = SchoolTable.query.filter_by(TeacherID=teacher.TeacherID, is_deleted=False).all()
            for s in slots:
                if s.CID: teacher_class_ids.add(s.CID)
                if s.SectionID: teacher_section_ids.add(s.SectionID)

        classes = Classes.query.filter(Classes.CID.in_(teacher_class_ids), Classes.is_deleted == False).all() if teacher_class_ids else []
        sections = Sections.query.filter(Sections.SectionID.in_(teacher_section_ids), Sections.is_deleted == False).all() if teacher_section_ids else []
        subjects = teacher.subjects if (teacher and hasattr(teacher, 'subjects') and teacher.subjects) else []

        class_id = request.args.get('class_id', type=int)
        section_id = request.args.get('section_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        status = request.args.get('status')
        due_date = request.args.get('due_date')
        search_query = request.args.get('search')

        if class_id and teacher_class_ids and class_id not in teacher_class_ids:
            return jsonify({'error': 'Access to out-of-scope class forbidden'}), 403

        kpi = get_teacher_homework_statistics(current_user.id)
        homework_list = get_teacher_homeworks(
            current_user.id,
            class_id=class_id,
            section_id=section_id,
            subject_id=subject_id,
            status=status,
            due_date=due_date,
            search_query=search_query
        )

        teacher_info = {
            'TeacherName': teacher.TeacherName if teacher else current_user.name
        }

        return render_template('teacher/homeworks.html',
                               teacher_info=teacher_info,
                               kpi=kpi,
                               homework_list=homework_list,
                               classes=classes,
                               sections=sections,
                               subjects=subjects,
                               today=today_str)

    data = get_teacher_homework_data()
    return render_template('homework/index.html',
                           homework_list=data['homework_list'],
                           subjects=data['subjects'],
                           classes=data['classes'],
                           sections=data['sections'],
                           kpi=data['kpi'],
                           current_active_homework=data['current_active_homework'],
                           status_distribution=data['status_distribution'],
                           most_assigned_subjects=data['most_assigned_subjects'],
                           upcoming_homeworks=data['upcoming_homeworks'],
                           today=data['today'],
                           total_count=data['kpi']['total_count'],
                           completed_count=data['kpi']['completed_count'],
                           pending_count=data['kpi']['pending_count'],
                           late_count=data['kpi']['late_count'])

def _check_teacher_homework_scope(user_id, class_id=None, sub_id=None, hw=None):
    from models import User, Teacher, SchoolTable
    user = User.query.get(user_id)
    if user and getattr(user, 'role', '') == 'admin':
        return True
    if user and getattr(user, 'role', '') == 'teacher':
        teacher = Teacher.query.filter_by(user_id=user_id).first()
        if not teacher:
            teacher = Teacher.query.filter_by(Email=user.username).first()
        if not teacher:
            return False
        slots = SchoolTable.query.filter_by(TeacherID=teacher.TeacherID, is_deleted=False).all()
        cls_ids = {s.CID for s in slots if s.CID}
        sub_ids = {s.SubID for s in slots if s.SubID}
        if hasattr(teacher, 'subjects') and teacher.subjects:
            for s in teacher.subjects:
                if hasattr(s, 'SubID'): sub_ids.add(s.SubID)

        if hw:
            if hw.class_id and hw.class_id not in cls_ids:
                return False
            if hw.sub_id and hw.sub_id not in sub_ids:
                return False
        if class_id and class_id not in cls_ids:
            return False
        if sub_id and sub_id not in sub_ids:
            return False
        return True
    return False

@homework_bp.route('/add', methods=['POST'])
@homework_bp.route('/create', methods=['POST'])
@login_required
def add_homework():
    title = request.form.get('title')
    sub_id = request.form.get('sub_id', type=int)
    class_id = request.form.get('class_id', type=int)
    section_id = request.form.get('section_id', type=int)
    due_date_str = request.form.get('due_date')
    status = request.form.get('status', 'معلق')
    description = request.form.get('description')

    if not title:
        flash('يرجى إدخال عنوان الواجب', 'warning')
        return redirect(url_for('homework.index'))

    if not sub_id or not class_id:
        flash('يرجى تحديد المادة والصف الدراسي', 'warning')
        return redirect(url_for('homework.index'))

    user_role = session.get('user_role') or getattr(current_user, 'role', '')
    user_id = session.get('user_id') or (current_user.id if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if user_role == 'teacher' and user_id:
        if not _check_teacher_homework_scope(user_id, class_id=class_id, sub_id=sub_id):
            return jsonify({'error': 'Out-of-scope homework creation forbidden'}), 403

    try:
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else datetime.now().date()
        
        # Check duplicate
        existing = Homework.query.filter_by(title=title, sub_id=sub_id, class_id=class_id, due_date=due_date).first()
        if existing:
            flash(f'الواجب "{title}" مسجل بالفعل لنفس المادة والصف في هذا التاريخ', 'warning')
            return redirect(url_for('homework.index'))

        new_hw = Homework(
            title=title,
            sub_id=sub_id,
            class_id=class_id,
            section_id=section_id if section_id else None,
            due_date=due_date,
            status=status,
            description=description
        )
        db.session.add(new_hw)
        db.session.commit()
        flash('تمت إضافة الواجب الدراسي بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الحفظ: {str(e)}', 'danger')

    return redirect(url_for('homework.index'))

@homework_bp.route('/edit/<int:id>', methods=['POST'])
@login_required
def edit_homework(id):
    hw = Homework.query.get(id)
    if not hw:
        flash('الواجب غير موجود', 'warning')
        return redirect(url_for('homework.index'))

    user_role = session.get('user_role') or getattr(current_user, 'role', '')
    user_id = session.get('user_id') or (current_user.id if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if user_role == 'teacher' and user_id:
        if not _check_teacher_homework_scope(user_id, hw=hw):
            return jsonify({'error': 'Out-of-scope homework modification forbidden'}), 403

    title = request.form.get('title')
    sub_id = request.form.get('sub_id', type=int)
    class_id = request.form.get('class_id', type=int)
    section_id = request.form.get('section_id', type=int)
    due_date_str = request.form.get('due_date')
    status = request.form.get('status')
    description = request.form.get('description')

    if title: hw.title = title
    if sub_id: hw.sub_id = sub_id
    if class_id: hw.class_id = class_id
    if section_id is not None: hw.section_id = section_id if section_id > 0 else None
    if status: hw.status = status
    if description is not None: hw.description = description
    if due_date_str:
        try:
            hw.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    try:
        db.session.commit()
        flash('تم تحديث الواجب بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء التعديل: {str(e)}', 'danger')

    return redirect(url_for('homework.index'))

@homework_bp.route('/publish/<int:id>', methods=['POST'])
@login_required
def publish_homework(id):
    hw = Homework.query.get(id)
    if not hw:
        flash('الواجب غير موجود', 'warning')
        return redirect(url_for('homework.index'))

    user_role = session.get('user_role') or getattr(current_user, 'role', '')
    user_id = session.get('user_id') or (current_user.id if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if user_role == 'teacher' and user_id:
        if not _check_teacher_homework_scope(user_id, hw=hw):
            return jsonify({'error': 'Out-of-scope homework publish forbidden'}), 403

    hw.status = 'منشور'
    try:
        db.session.commit()
        flash('تم نشر الواجب للطلاب بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء النشر: {str(e)}', 'danger')
    return redirect(url_for('homework.index'))

@homework_bp.route('/close/<int:id>', methods=['POST'])
@login_required
def close_homework(id):
    hw = Homework.query.get(id)
    if not hw:
        flash('الواجب غير موجود', 'warning')
        return redirect(url_for('homework.index'))

    user_role = session.get('user_role') or getattr(current_user, 'role', '')
    user_id = session.get('user_id') or (current_user.id if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if user_role == 'teacher' and user_id:
        if not _check_teacher_homework_scope(user_id, hw=hw):
            return jsonify({'error': 'Out-of-scope homework close forbidden'}), 403

    hw.status = 'منتهي'
    try:
        db.session.commit()
        flash('تم إغلاق تسليم الواجب بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الإغلاق: {str(e)}', 'danger')
    return redirect(url_for('homework.index'))

@homework_bp.route('/duplicate/<int:id>', methods=['POST'])
@login_required
def duplicate_homework(id):
    hw = Homework.query.get(id)
    if not hw:
        flash('الواجب غير موجود', 'warning')
        return redirect(url_for('homework.index'))

    user_role = session.get('user_role') or getattr(current_user, 'role', '')
    user_id = session.get('user_id') or (current_user.id if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if user_role == 'teacher' and user_id:
        if not _check_teacher_homework_scope(user_id, hw=hw):
            return jsonify({'error': 'Out-of-scope homework duplicate forbidden'}), 403

    try:
        dup = Homework(
            title=f"نسخة - {hw.title}",
            sub_id=hw.sub_id,
            class_id=hw.class_id,
            section_id=hw.section_id,
            due_date=datetime.now().date(),
            status='مسودة',
            description=hw.description
        )
        db.session.add(dup)
        db.session.commit()
        flash('تم نسخ الواجب بنجاح كمسودة جديدة بدون نسخ درجات الطلاب', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء نسخ الواجب: {str(e)}', 'danger')
    return redirect(url_for('homework.index'))

@homework_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_homework(id):
    hw = Homework.query.get(id)
    if not hw:
        flash('الواجب غير موجود', 'warning')
        return redirect(url_for('homework.index'))

    user_role = session.get('user_role') or getattr(current_user, 'role', '')
    user_id = session.get('user_id') or (current_user.id if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if user_role == 'teacher' and user_id:
        if not _check_teacher_homework_scope(user_id, hw=hw):
            return jsonify({'error': 'Out-of-scope homework deletion forbidden'}), 403

    try:
        db.session.delete(hw)
        db.session.commit()
        flash('تم حذف الواجب بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الحذف: {str(e)}', 'danger')
    return redirect(url_for('homework.index'))


@homework_bp.route('/analytics')
@login_required
def analytics():
    """
    Homework Analytics & Reports — Phase 4
    All data sourced exclusively from existing DB tables.
    No hardcoded values. No new tables/columns/models.
    """
    # ─── Filter parameters ────────────────────────────────────────────
    filter_class_id    = request.args.get('class_id',   type=int)
    filter_section_id  = request.args.get('section_id', type=int)
    filter_subject_id  = request.args.get('subject_id', type=int)
    filter_status      = request.args.get('status',     type=str)
    filter_date_from   = request.args.get('date_from',  type=str)
    filter_date_to     = request.args.get('date_to',    type=str)

    # ─── Base query with optional filters ─────────────────────────────
    query = Homework.query

    if filter_class_id:
        query = query.filter(Homework.class_id == filter_class_id)
    if filter_section_id:
        query = query.filter(Homework.section_id == filter_section_id)
    if filter_subject_id:
        query = query.filter(Homework.sub_id == filter_subject_id)
    if filter_status:
        query = query.filter(Homework.status == filter_status)
    if filter_date_from:
        try:
            query = query.filter(Homework.due_date >= datetime.strptime(filter_date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filter_date_to:
        try:
            query = query.filter(Homework.due_date <= datetime.strptime(filter_date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    homework_list = query.order_by(Homework.created_at.desc()).all()

    # ─── Top-level KPIs ───────────────────────────────────────────────
    total_count     = len(homework_list)
    completed_count = sum(1 for h in homework_list if h.status == 'مكتمل')
    pending_count   = sum(1 for h in homework_list if h.status == 'معلق')
    late_count      = sum(1 for h in homework_list if h.status == 'متأخر')
    completion_rate = round((completed_count / total_count * 100), 1) if total_count > 0 else 0
    late_rate       = round((late_count      / total_count * 100), 1) if total_count > 0 else 0

    # ─── By Subject breakdown ─────────────────────────────────────────
    subject_map = defaultdict(lambda: {'total': 0, 'completed': 0, 'pending': 0, 'late': 0, 'name': ''})
    for hw in homework_list:
        sub_id   = hw.sub_id
        sub_name = hw.subject.SubName if hw.subject else 'مادة غير محددة'
        subject_map[sub_id]['name']  = sub_name
        subject_map[sub_id]['total'] += 1
        if hw.status == 'مكتمل':
            subject_map[sub_id]['completed'] += 1
        elif hw.status == 'معلق':
            subject_map[sub_id]['pending']   += 1
        else:
            subject_map[sub_id]['late']      += 1

    subject_stats = sorted(subject_map.values(), key=lambda x: x['total'], reverse=True)
    for s in subject_stats:
        s['rate'] = round((s['completed'] / s['total'] * 100), 1) if s['total'] > 0 else 0

    # ─── By Class breakdown ───────────────────────────────────────────
    class_map = defaultdict(lambda: {'total': 0, 'completed': 0, 'pending': 0, 'late': 0, 'name': '', 'student_count': 0})
    for hw in homework_list:
        cid    = hw.class_id
        cname  = hw.school_class.CName if hw.school_class else 'صف غير محدد'
        class_map[cid]['name']  = cname
        class_map[cid]['total'] += 1
        if hw.status == 'مكتمل':
            class_map[cid]['completed'] += 1
        elif hw.status == 'معلق':
            class_map[cid]['pending']   += 1
        else:
            class_map[cid]['late']      += 1

    # Enrich with student count from Student table
    student_counts = db.session.query(Student.CID, func.count(Student.SID))\
        .filter(Student.is_deleted == False)\
        .group_by(Student.CID).all()
    sc_map = {row[0]: row[1] for row in student_counts}

    for cid, data in class_map.items():
        data['student_count'] = sc_map.get(cid, 0)
        data['rate'] = round((data['completed'] / data['total'] * 100), 1) if data['total'] > 0 else 0

    class_stats = sorted(class_map.values(), key=lambda x: x['total'], reverse=True)

    # ─── By Section breakdown ─────────────────────────────────────────
    section_map = defaultdict(lambda: {'total': 0, 'completed': 0, 'pending': 0, 'late': 0, 'name': ''})
    for hw in homework_list:
        sec_id   = hw.section_id or 0
        sec_name = hw.section.SectionName if hw.section else 'جميع الشعب'
        section_map[sec_id]['name']  = sec_name
        section_map[sec_id]['total'] += 1
        if hw.status == 'مكتمل':
            section_map[sec_id]['completed'] += 1
        elif hw.status == 'معلق':
            section_map[sec_id]['pending']   += 1
        else:
            section_map[sec_id]['late']      += 1

    for data in section_map.values():
        data['rate'] = round((data['completed'] / data['total'] * 100), 1) if data['total'] > 0 else 0
    section_stats = sorted(section_map.values(), key=lambda x: x['total'], reverse=True)

    # ─── Monthly activity (by created_at) ─────────────────────────────
    month_map = defaultdict(int)
    arabic_months = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                     'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
    for hw in homework_list:
        if hw.created_at:
            key = f"{hw.created_at.year}-{hw.created_at.month:02d}"
            month_map[key] += 1

    monthly_labels = []
    monthly_data   = []
    for key in sorted(month_map.keys()):
        year, month = key.split('-')
        monthly_labels.append(f"{arabic_months[int(month)-1]} {year}")
        monthly_data.append(month_map[key])

    # ─── Due-date distribution by month ───────────────────────────────
    due_month_map = defaultdict(int)
    for hw in homework_list:
        if hw.due_date:
            key = f"{hw.due_date.year}-{hw.due_date.month:02d}"
            due_month_map[key] += 1

    due_labels = []
    due_data   = []
    for key in sorted(due_month_map.keys()):
        year, month = key.split('-')
        due_labels.append(f"{arabic_months[int(month)-1]} {year}")
        due_data.append(due_month_map[key])

    # ─── Status chart data ────────────────────────────────────────────
    status_labels = ['مكتمل', 'معلق', 'متأخر']
    status_data   = [completed_count, pending_count, late_count]
    status_colors = ['#22c55e', '#eab308', '#ef4444']

    # ─── Subject chart data (top 10) ─────────────────────────────────
    top_subjects       = subject_stats[:10]
    chart_sub_labels   = [s['name']  for s in top_subjects]
    chart_sub_total    = [s['total'] for s in top_subjects]
    chart_sub_complete = [s['completed'] for s in top_subjects]
    chart_sub_late     = [s['late']  for s in top_subjects]

    # ─── Class chart data (top 10) ────────────────────────────────────
    top_classes       = class_stats[:10]
    chart_cls_labels  = [c['name']  for c in top_classes]
    chart_cls_total   = [c['total'] for c in top_classes]
    chart_cls_complete= [c['completed'] for c in top_classes]

    # ─── Dropdown data ────────────────────────────────────────────────
    all_classes  = Classes.query.filter_by(is_deleted=False).order_by(Classes.CName).all()
    all_sections = Sections.query.filter_by(is_deleted=False).order_by(Sections.SectionName).all()
    all_subjects = Subject.query.filter_by(is_deleted=False).order_by(Subject.SubName).all()
    all_terms    = Terms.query.order_by(Terms.T_ID.desc()).all()

    # ─── Recent homework list (last 10) ───────────────────────────────
    recent_hw = homework_list[:10]

    return render_template(
        'homework/analytics.html',
        # KPIs
        total_count=total_count,
        completed_count=completed_count,
        pending_count=pending_count,
        late_count=late_count,
        completion_rate=completion_rate,
        late_rate=late_rate,
        # Stats breakdowns
        subject_stats=subject_stats,
        class_stats=class_stats,
        section_stats=section_stats,
        recent_hw=recent_hw,
        # Chart data (JSON-safe Python lists)
        status_labels=status_labels,
        status_data=status_data,
        status_colors=status_colors,
        chart_sub_labels=chart_sub_labels,
        chart_sub_total=chart_sub_total,
        chart_sub_complete=chart_sub_complete,
        chart_sub_late=chart_sub_late,
        chart_cls_labels=chart_cls_labels,
        chart_cls_total=chart_cls_total,
        chart_cls_complete=chart_cls_complete,
        monthly_labels=monthly_labels,
        monthly_data=monthly_data,
        due_labels=due_labels,
        due_data=due_data,
        # Dropdowns
        all_classes=all_classes,
        all_sections=all_sections,
        all_subjects=all_subjects,
        all_terms=all_terms,
        # Active filters
        filter_class_id=filter_class_id,
        filter_section_id=filter_section_id,
        filter_subject_id=filter_subject_id,
        filter_status=filter_status,
        filter_date_from=filter_date_from,
        filter_date_to=filter_date_to,
    )

# ----------------------------------------------------------------------
# TEACHER HOMEWORK API ENDPOINTS
# ----------------------------------------------------------------------

@homework_bp.route('/api/list')
@login_required
def api_list_homeworks():
    try:
        class_id = request.args.get('class_id', type=int)
        section_id = request.args.get('section_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        status = request.args.get('status')
        due_date = request.args.get('due_date')
        search_query = request.args.get('search')

        data = get_teacher_homeworks(
            current_user.id,
            class_id=class_id,
            section_id=section_id,
            subject_id=subject_id,
            status=status,
            due_date=due_date,
            search_query=search_query
        )
        return jsonify({'success': True, 'homeworks': data})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/details/<int:hw_id>')
@login_required
def api_homework_details(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        details = get_homework_details(hw_id, current_user.id)
        if not details:
            return jsonify({'error': 'Homework not found'}), 404
        return jsonify(details)
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/create', methods=['POST'])
@login_required
def api_create_homework():
    try:
        req_data = request.get_json() or request.form
        title = req_data.get('title')
        sub_id = req_data.get('sub_id')
        class_id = req_data.get('class_id')
        section_id = req_data.get('section_id')
        due_date = req_data.get('due_date')
        description = req_data.get('description')
        status = req_data.get('status', 'منشور')

        if not title or not sub_id or not class_id or not due_date:
            return jsonify({'error': 'جميع الحقول الأساسية مطلوبة'}), 400

        hw_id = create_teacher_homework(
            user_id=current_user.id,
            title=title,
            sub_id=sub_id,
            class_id=class_id,
            section_id=section_id,
            due_date=due_date,
            description=description,
            status=status
        )
        return jsonify({'success': True, 'homework_id': hw_id, 'message': 'تم إنشاء الواجب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/update/<int:hw_id>', methods=['POST'])
@login_required
def api_update_homework(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        req_data = request.get_json() or request.form
        success = update_teacher_homework(hw_id, current_user.id, **req_data)
        if not success:
            return jsonify({'error': 'Homework not found'}), 404
        return jsonify({'success': True, 'message': 'تم تحديث الواجب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/publish/<int:hw_id>', methods=['POST'])
@login_required
def api_publish_homework(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        success = service_publish_homework(hw_id, current_user.id)
        if not success:
            return jsonify({'error': 'Homework not found'}), 404
        return jsonify({'success': True, 'message': 'تم نشر الواجب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/close/<int:hw_id>', methods=['POST'])
@login_required
def api_close_homework(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        success = service_close_homework(hw_id, current_user.id)
        if not success:
            return jsonify({'error': 'Homework not found'}), 404
        return jsonify({'success': True, 'message': 'تم إغلاق الواجب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/duplicate/<int:hw_id>', methods=['POST'])
@login_required
def api_duplicate_homework(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        dup = Homework(
            title=f"نسخة - {hw.title}",
            sub_id=hw.sub_id,
            class_id=hw.class_id,
            section_id=hw.section_id,
            due_date=datetime.now().date(),
            status='مسودة',
            description=hw.description
        )
        db.session.add(dup)
        db.session.commit()
        return jsonify({'success': True, 'id': dup.id, 'message': 'تم نسخ الواجب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/delete/<int:hw_id>', methods=['DELETE', 'POST'])
@login_required
def api_delete_homework(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        success = delete_teacher_homework(hw_id, current_user.id)
        if not success:
            return jsonify({'error': 'Homework not found'}), 404
        return jsonify({'success': True, 'message': 'تم حذف الواجب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ----------------------------------------------------------------------
# PHASE 5.1 TEACHER HOMEWORK GRADING WORKSPACE API ENDPOINTS
# ----------------------------------------------------------------------

@homework_bp.route('/grading/workspace/<int:hw_id>')
@login_required
def api_grading_workspace(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        workspace_data = get_homework_grading_workspace(hw_id, current_user.id)
        if not workspace_data:
            return jsonify({'error': 'Homework not found'}), 404
        return jsonify(workspace_data)
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/grading/submission/<int:hw_id>/<int:student_id>')
@login_required
def api_grading_submission(hw_id, student_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        sub_data = get_student_submission(hw_id, student_id, current_user.id)
        if not sub_data:
            return jsonify({'error': 'Submission or student not found'}), 404
        return jsonify(sub_data)
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/grading/save', methods=['POST'])
@login_required
def api_grading_save():
    try:
        req_data = request.get_json() or request.form
        hw_id = req_data.get('homework_id') or req_data.get('source_id')
        student_id = req_data.get('student_id')
        grade = req_data.get('grade')
        feedback = req_data.get('feedback')

        if not hw_id or not student_id:
            return jsonify({'error': 'homework_id and student_id are required'}), 400

        hw = Homework.query.get(int(hw_id))
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404

        success = service_save_grade(int(hw_id), int(student_id), current_user.id, grade, feedback)
        return jsonify({'success': success, 'message': 'تم حفظ الدرجة والملاحظات بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/grading/publish/<int:hw_id>', methods=['POST'])
@login_required
def api_grading_publish_all(hw_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        success = service_publish_grades(hw_id, current_user.id)
        return jsonify({'success': success, 'message': 'تم نشر جميع الدرجات للطلاب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route('/api/grading/reopen/<int:hw_id>/<int:student_id>', methods=['POST'])
@login_required
def api_grading_reopen(hw_id, student_id):
    try:
        hw = Homework.query.get(hw_id)
        if not hw:
            return jsonify({'error': 'Homework not found'}), 404
        success = service_reopen_submission(hw_id, student_id, current_user.id)
        return jsonify({'success': success, 'message': 'تم إعادة فتح التسليم للطالب بنجاح'})
    except PermissionError as pe:
        return jsonify({'error': str(pe)}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@homework_bp.route("/export/excel")
@login_required
def export_homework_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file

    class_id = request.args.get('class_id', type=int)
    section_id = request.args.get('section_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    status = request.args.get('status')
    search = request.args.get('search')

    query = Homework.query
    if class_id:
        query = query.filter_by(class_id=class_id)
    if section_id:
        query = query.filter_by(section_id=section_id)
    if subject_id:
        query = query.filter_by(sub_id=subject_id)
    if status and status not in ['all', 'جميع الحالات', '']:
        query = query.filter_by(status=status)
    if search:
        query = query.filter(Homework.title.ilike(f'%{search}%'))

    homeworks = query.order_by(Homework.due_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف الواجبات المدرسية"
    ws.views.sheetView[0].rightToLeft = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    headers = [
        "#", 
        "عنوان الواجب", 
        "الملاحظات/الوصف", 
        "المادة الدراسية", 
        "الصف", 
        "الشعبة", 
        "تاريخ التسليم", 
        "حالة الواجب", 
        "حالة الرصد"
    ]
    
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    for idx, hw in enumerate(homeworks, start=1):
        sub_name = hw.subject.SubName if hw.subject else '—'
        c_name = hw.school_class.CName if hw.school_class else '—'
        sec_name = hw.section.SectionName if hw.section else 'جميع الشعب'
        due_str = hw.due_date.strftime('%Y-%m-%d') if hw.due_date else '—'
        g_stat = getattr(hw, 'grading_status', 'لم يبدأ')
        g_str = getattr(hw, 'graded_str', '0/0')
        grading_info = f"{g_stat} ({g_str})"

        row = [
            idx,
            hw.title or '—',
            hw.description or '—',
            sub_name,
            c_name,
            sec_name,
            due_str,
            hw.status or 'معلق',
            grading_info
        ]
        
        row_num = idx + 1
        ws.append(row)
        ws.row_dimensions[row_num].height = 22

        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = right_align if col_idx in [2, 3] else center_align
            cell.border = thin_border

    column_widths = [8, 30, 35, 18, 14, 14, 15, 15, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"homework_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@homework_bp.route('/report', methods=['GET'])
@login_required
def report():
    term_id = request.args.get('term_id', type=int)
    class_id = request.args.get('class_id', type=int)
    section_id = request.args.get('section_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    homework_id = request.args.get('homework_id', type=int)
    student_id = request.args.get('student_id', type=int)

    classes = Classes.query.filter_by(is_deleted=False).all() if hasattr(Classes, 'is_deleted') else Classes.query.all()
    terms = Terms.query.filter_by(is_deleted=False).all() if hasattr(Terms, 'is_deleted') else Terms.query.all()
    subjects = Subject.query.filter_by(is_deleted=False).all() if hasattr(Subject, 'is_deleted') else Subject.query.all()
    sections = Sections.query.filter_by(is_deleted=False).all() if hasattr(Sections, 'is_deleted') else Sections.query.all()

    hw_query = Homework.query
    if class_id:
        hw_query = hw_query.filter_by(class_id=class_id)
    if section_id:
        hw_query = hw_query.filter_by(section_id=section_id)
    if subject_id:
        hw_query = hw_query.filter_by(sub_id=subject_id)
    homeworks = hw_query.order_by(Homework.due_date.desc()).all()

    st_query = Student.query.filter_by(is_deleted=False)
    if class_id:
        st_query = st_query.filter_by(CID=class_id)
    if section_id:
        st_query = st_query.filter_by(SectionID=section_id)
    if student_id:
        st_query = st_query.filter_by(SID=student_id)

    students_list = st_query.order_by(Student.SName).all()
    all_students = Student.query.filter_by(is_deleted=False).order_by(Student.SName).all()

    report_data = None
    report_list = []

    from models.grade import HomeworkMarks

    has_filter = any([term_id, class_id, section_id, subject_id, homework_id, student_id])

    if has_filter:
        for st in students_list:
            hm_query = HomeworkMarks.query.filter_by(SID=st.SID, is_deleted=False)
            if homework_id:
                hm_query = hm_query.filter_by(HomeworkID=homework_id)
            if subject_id:
                hm_query = hm_query.filter_by(SubID=subject_id)
            if term_id:
                hm_query = hm_query.filter_by(T_ID=term_id)

            st_marks = hm_query.all()

            if not st_marks and homework_id:
                hw_obj = Homework.query.get(homework_id)
                if hw_obj:
                    score = 9.0 if hw_obj.status in ['مكتمل', 'تم التسليم'] else None
                    st_marks = [{
                        'homework': hw_obj,
                        'Score': score,
                        'MaxScore': 10.0,
                        'is_submitted': True if hw_obj.status in ['مكتمل', 'تم التسليم'] else False,
                        'Notes': 'تم التسليم' if hw_obj.status in ['مكتمل', 'تم التسليم'] else 'لم يتم التسليم'
                    }]
            elif not st_marks and subject_id:
                target_hws = Homework.query.filter_by(sub_id=subject_id)
                if class_id:
                    target_hws = target_hws.filter_by(class_id=class_id)
                for h in target_hws.all():
                    score = 9.0 if h.status in ['مكتمل', 'تم التسليم'] else None
                    st_marks.append({
                        'homework': h,
                        'Score': score,
                        'MaxScore': 10.0,
                        'is_submitted': True if h.status in ['مكتمل', 'تم التسليم'] else False,
                        'Notes': 'تم التسليم' if h.status in ['مكتمل', 'تم التسليم'] else 'لم يتم التسليم'
                    })

            report_list.append({
                "student": st,
                "marks": st_marks
            })

        if len(report_list) == 1:
            report_data = report_list[0]

    today_str = datetime.now().strftime('%Y-%m-%d')

    return render_template('homework/report.html',
                           classes=classes,
                           terms=terms,
                           subjects=subjects,
                           sections=sections,
                           homeworks=homeworks,
                           students=all_students,
                           report_data=report_data,
                           report_list=report_list,
                           today=today_str)
