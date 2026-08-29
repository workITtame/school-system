from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from flask_login import login_required, current_user
from models import db, Student, Subject, Classes, Sections, Terms, TypeExams, DetailMarks, Marks, Teacher

grades_bp = Blueprint('grades', __name__, url_prefix='/grades')
grades_legacy_bp = Blueprint('grades_legacy', __name__, url_prefix='/grades_legacy')

@grades_bp.route('/', methods=['GET'])
@grades_legacy_bp.route('/', methods=['GET'])
def index():
    if not current_user.is_authenticated and 'user_id' not in session:
        return redirect(url_for('auth.login'))
    user_role = getattr(current_user, 'role', '').strip("'") if current_user and hasattr(current_user, 'role') else None
    if user_role == 'teacher':
        return redirect(url_for('gradebook.index'))
    return redirect(url_for('grades.manage_grades'))

@grades_bp.route('/manage', methods=['GET'])
@grades_legacy_bp.route('/manage', methods=['GET'])
def manage_grades():
    if not current_user.is_authenticated and 'user_id' not in session:
        return redirect(url_for('auth.login'))
    user_role = getattr(current_user, 'role', '').strip("'") if current_user and hasattr(current_user, 'role') else None
    if user_role == 'teacher':
        return redirect(url_for('gradebook.index'))
        
    from models.academic import ExamSchedule
    total_students = Student.query.filter_by(is_deleted=False).count()
    total_exams = ExamSchedule.query.filter_by(is_deleted=False).count()
    total_subjects = Subject.query.filter_by(is_deleted=False).count()
    total_classes = Classes.query.filter_by(is_deleted=False).count()
    
    valid_marks = Marks.query.join(Student, Marks.SID == Student.SID).filter(Student.is_deleted == False).all()
    total_marks_count = len(valid_marks)
    
    scores = [float(m.Score) for m in valid_marks if m.Score is not None]
    from services.grade_calculation_service import calculate_exam_average, is_passing
    if scores:
        avg_score = calculate_exam_average(scores) or 0.0
        max_score = max(scores)
        min_score = min(scores)
        pass_count = sum(1 for s in scores if is_passing(s))
        fail_count = sum(1 for s in scores if not is_passing(s))
        pass_rate = round((pass_count / len(scores)) * 100, 1)
        fail_rate = round((fail_count / len(scores)) * 100, 1)
        rating_label = 'ممتاز' if avg_score >= 90 else ('جيد جداً' if avg_score >= 80 else ('جيد' if avg_score >= 70 else ('مقبول' if is_passing(avg_score) else 'ضعيف')))
    else:
        avg_score, max_score, min_score, pass_count, fail_count, pass_rate, fail_rate = 0.0, 0.0, 0.0, 0, 0, 0.0, 0.0
        rating_label = '—'
    
    from models.academic import Sections
    terms = Terms.query.filter_by(is_deleted=False).all()
    classes = Classes.query.filter_by(is_deleted=False).all()
    exams = TypeExams.query.filter_by(is_deleted=False).all()
    subjects = Subject.query.filter_by(is_deleted=False).all()
    sections = Sections.query.filter_by(is_deleted=False).all() if hasattr(Sections, 'is_deleted') else Sections.query.all()
    
    active_sched = ExamSchedule.query.filter_by(is_deleted=False).order_by(ExamSchedule.ScheduleID.desc()).first()

    # Dynamic Subject Averages
    subject_stats = []
    for sub in subjects:
        sub_scores = [float(m.Score) for m in valid_marks if m.SubID == sub.SubID and m.Score is not None]
        sub_avg = round(sum(sub_scores) / len(sub_scores), 1) if sub_scores else 0.0
        subject_stats.append({"name": sub.SubName, "average": sub_avg})

    sorted_sub_with_scores = [s for s in subject_stats if s["average"] > 0]
    if sorted_sub_with_scores:
        sorted_sub_with_scores.sort(key=lambda x: x["average"], reverse=True)
        best_subject_str = f"{sorted_sub_with_scores[0]['name']} (متوسط {sorted_sub_with_scores[0]['average']}%)"
        hardest_subject_str = f"{sorted_sub_with_scores[-1]['name']} (متوسط {sorted_sub_with_scores[-1]['average']}%)"
    else:
        best_subject_str = "لا توجد درجات مسجلة"
        hardest_subject_str = "لا توجد درجات مسجلة"

    # Dynamic Exam Trends
    exam_trends = []
    for ex in exams:
        ex_scores = [float(m.Score) for m in valid_marks if m.ExamID == ex.ExamID and m.Score is not None]
        ex_avg = round(sum(ex_scores) / len(ex_scores), 1) if ex_scores else 0.0
        exam_trends.append({"name": ex.ExamName, "average": ex_avg})

    sorted_exam_with_scores = [e for e in exam_trends if e["average"] > 0]
    if sorted_exam_with_scores:
        sorted_exam_with_scores.sort(key=lambda x: x["average"], reverse=True)
        highest_exam_str = f"{sorted_exam_with_scores[0]['name']} (متوسط {sorted_exam_with_scores[0]['average']}%)"
        lowest_exam_str = f"{sorted_exam_with_scores[-1]['name']} (متوسط {sorted_exam_with_scores[-1]['average']}%)"
    else:
        highest_exam_str = "لا توجد بيانات"
        lowest_exam_str = "لا توجد بيانات"

    # Student Rankings & AI Insights
    student_avg_map = {}
    for m in valid_marks:
        if m.Score is not None:
            student_avg_map.setdefault(m.SID, []).append(float(m.Score))

    top_students_list = []
    struggling_students_list = []
    for sid, score_list in student_avg_map.items():
        st_obj = Student.query.get(sid)
        if st_obj and not getattr(st_obj, 'is_deleted', False):
            st_name = getattr(st_obj, 'SName', None) or getattr(st_obj, 'StudentName', 'طالب')
            st_avg = round(sum(score_list) / len(score_list), 1)
            if st_avg >= 90:
                top_students_list.append(st_name)
            elif st_avg < 60:
                struggling_students_list.append(st_name)

    top_students_str = "، ".join(top_students_list[:3]) if top_students_list else "لا يوجد طلاب متفوقون ممتاز"
    struggling_students_str = "، ".join(struggling_students_list[:3]) if struggling_students_list else "لا يوجد طلاب متعثرون"

    stats = {
        "total_students": total_students,
        "total_exams": total_exams,
        "total_subjects": total_subjects,
        "total_classes": total_classes,
        "total_marks_count": total_marks_count,
        "avg_score": avg_score,
        "max_score": max_score,
        "min_score": min_score,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "pass_rate": pass_rate,
        "fail_rate": fail_rate,
        "rating_label": rating_label,
        "active_exam_name": active_sched.ExamName if active_sched else (exams[0].ExamName if exams else 'اختبار شهري'),
        "active_subject_name": active_sched.subject.SubName if active_sched and active_sched.subject else (subjects[0].SubName if subjects else 'المادة العامة'),
        "active_class_name": active_sched.school_class.CName if active_sched and active_sched.school_class else (classes[0].CName if classes else 'الصف العام'),
        "active_section_name": active_sched.section.SectionName if active_sched and active_sched.section else (sections[0].SectionName if sections else 'جميع الشعب'),
        "best_subject": best_subject_str,
        "hardest_subject": hardest_subject_str,
        "highest_exam": highest_exam_str,
        "lowest_exam": lowest_exam_str,
        "top_students": top_students_str,
        "struggling_students": struggling_students_str,
        "subject_stats": subject_stats,
        "exam_trends": exam_trends
    }

    all_students = Student.query.filter_by(is_deleted=False).order_by(Student.SName).all()

    return render_template('grades/manage.html', 
                           stats=stats, 
                           terms=terms, 
                           classes=classes, 
                           exams=exams, 
                           subjects=subjects,
                           sections=sections,
                           all_students=all_students)

@grades_bp.route('/report', methods=['GET'])
@grades_legacy_bp.route('/report', methods=['GET'])
def student_report_page():
    if 'user_id' not in session and not (current_user and current_user.is_authenticated):
        return redirect(url_for('auth.login'))
        
    from models.academic import Sections
    classes = Classes.query.filter_by(is_deleted=False).all() if hasattr(Classes, 'is_deleted') else Classes.query.all()
    terms = Terms.query.filter_by(is_deleted=False).all() if hasattr(Terms, 'is_deleted') else Terms.query.all()
    exams = TypeExams.query.filter_by(is_deleted=False).all() if hasattr(TypeExams, 'is_deleted') else TypeExams.query.all()
    subjects = Subject.query.filter_by(is_deleted=False).all() if hasattr(Subject, 'is_deleted') else Subject.query.all()
    sections = Sections.query.filter_by(is_deleted=False).all() if hasattr(Sections, 'is_deleted') else Sections.query.all()
    
    student_id = request.args.get('student_id', type=int)
    term_id = request.args.get('term_id', type=int)
    class_id = request.args.get('class_id', type=int)
    section_id = request.args.get('section_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    exam_id = request.args.get('exam_id', type=int)
    
    st_query = Student.query.filter_by(is_deleted=False)
    if class_id:
        st_query = st_query.filter_by(CID=class_id)
    if section_id:
        st_query = st_query.filter_by(SectionID=section_id)
    if student_id:
        st_query = st_query.filter_by(SID=student_id)

    students_list = st_query.order_by(Student.SName).all()
    
    report_data = None
    report_list = []
    
    has_filter = any([term_id, class_id, section_id, subject_id, exam_id, student_id])

    if has_filter:
        for st in students_list:
            m_query = Marks.query.filter_by(SID=st.SID)
            if term_id:
                m_query = m_query.filter_by(T_ID=term_id)
            if exam_id:
                m_query = m_query.filter_by(ExamID=exam_id)
            if subject_id:
                m_query = m_query.filter_by(SubID=subject_id)
            
            st_marks = m_query.all()
            report_list.append({
                "student": st,
                "marks": st_marks
            })

        if len(report_list) == 1:
            report_data = report_list[0]

    all_students = Student.query.filter_by(is_deleted=False).order_by(Student.SName).all()

    return render_template('grades/student_report.html', 
                           classes=classes, 
                           terms=terms, 
                           exams=exams, 
                           subjects=subjects,
                           sections=sections,
                           students=all_students,
                           report_data=report_data,
                           report_list=report_list)

@grades_bp.route('/add_exam', methods=['POST'])
@grades_legacy_bp.route('/add_exam', methods=['POST'])
def add_exam():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    exam_name = request.form.get('exam_name') or request.form.get('ExamName')
    if exam_name:
        try:
            new_type = TypeExams(ExamName=exam_name)
            db.session.add(new_type)
            db.session.commit()
            flash('تم إضافة نوع الاختبار بنجاح', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ عند الإضافة: {e}', 'danger')
    return redirect(url_for('grades.manage_grades'))


@grades_bp.route('/export/excel', methods=['GET'])
@grades_legacy_bp.route('/export/excel', methods=['GET'])
def export_grades_excel():
    from io import BytesIO
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from models import Attendance

    if 'user_id' not in session and not (current_user and current_user.is_authenticated):
        return redirect(url_for('auth.login'))
        
    class_id = request.args.get('class_id', type=int)
    section_id = request.args.get('section_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    exam_id = request.args.get('exam_id', type=int)
    term_id = request.args.get('term_id', type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Student.query.filter_by(is_deleted=False)
    if class_id:
        query = query.filter_by(CID=class_id)
    if section_id:
        query = query.filter_by(SectionID=section_id)
        
    students = query.order_by(Student.SName).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف درجات الاختبارات"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "#", 
        "الرقم الأكاديمي", 
        "اسم الطالب", 
        "الصف الدراسي", 
        "الشعبة", 
        "المادة الدراسية", 
        "نوع الاختبار", 
        "الحضور والغياب", 
        "الدرجة المرصودة", 
        "النسبة المئوية", 
        "التقدير الأكاديمي", 
        "حالة الاعتماد"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    row_index = 1
    for st in students:
        cname = st.school_class.CName if st.school_class else "جميع الصفوف"
        sname = st.section.SectionName if st.section else "جميع الشعب"
        
        mark_q = Marks.query.filter_by(SID=st.SID)
        if subject_id: mark_q = mark_q.filter_by(SubID=subject_id)
        if exam_id: mark_q = mark_q.filter_by(ExamID=exam_id)
        if term_id: mark_q = mark_q.filter_by(T_ID=term_id)
        
        mark = mark_q.first()

        score = float(mark.Score) if mark and mark.Score is not None else None
        
        if status_filter == 'approved' and score is None:
            continue
        if status_filter == 'missing' and score is not None:
            continue

        if score is not None:
            score_str = f"{score}"
            percent_str = f"{score:.1f}%"
            if score >= 90:
                rating = "ممتاز"
            elif score >= 80:
                rating = "جيد جداً"
            elif score >= 70:
                rating = "جيد"
            elif score >= 60:
                rating = "مقبول"
            else:
                rating = "راسب"
            approval = "معتمد"
        else:
            score_str = "—"
            percent_str = "—"
            rating = "غير مدخل"
            approval = "لم ترصد"

        att = Attendance.query.filter_by(SID=st.SID).order_by(Attendance.Date.desc()).first()
        att_str = att.Status if att else "حاضر"

        sub_name = mark.subject.SubName if mark and mark.subject else "جميع المواد"
        exam_name = mark.exam.ExamName if mark and hasattr(mark, 'exam') and mark.exam else "جميع الاختبارات"

        row = [
            row_index,
            st.SID,
            st.SName if hasattr(st, 'SName') else st.StudentName,
            cname,
            sname,
            sub_name,
            exam_name,
            att_str,
            score_str,
            percent_str,
            rating,
            approval
        ]
        
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
        row_index += 1

    col_widths = {
        'A': 6, 'B': 18, 'C': 26, 'D': 16, 'E': 14, 
        'F': 18, 'G': 18, 'H': 14, 'I': 16, 'J': 16, 
        'K': 16, 'L': 16
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"كشف_درجات_الاختبارات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

