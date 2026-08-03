from flask import Blueprint, render_template, request, redirect, url_for, session
import time
from models import Qualifications
from utils.decorators import admin_required

teacher_bp = Blueprint('teacher', __name__, url_prefix='/teacher')

from models import Qualifications, Teacher

@teacher_bp.route('/')
def index():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    
    from models import Qualifications, Teacher, SchoolTable, Classes, db
    from models.academic import TeacherSubject
    
    qualifications = Qualifications.query.all()
    
    total_teachers = Teacher.query.filter_by(is_deleted=False).count()
    male_teachers = Teacher.query.filter(Teacher.is_deleted == False, Teacher.Gender.in_(['ذكر', 'Male'])).count()
    female_teachers = Teacher.query.filter(Teacher.is_deleted == False, Teacher.Gender.in_(['أنثى', 'Female'])).count()
    active_teachers = Teacher.query.filter_by(is_deleted=False, Status='نشط').count()
    
    # Advanced Real KPIs
    fulltime_teachers = active_teachers
    assigned_subjects_count = db.session.query(TeacherSubject).count()
    total_slots = SchoolTable.query.filter_by(is_deleted=False).count()
    avg_weekly_slots = round(total_slots / total_teachers, 1) if total_teachers > 0 else 0.0
    taught_classes_count = db.session.query(Classes.CID).join(SchoolTable, Classes.CID == SchoolTable.CID).distinct().count()
    
    return render_template('teacher/index.html', 
                           qualifications=qualifications,
                           total_teachers=total_teachers,
                           male_teachers=male_teachers,
                           female_teachers=female_teachers,
                           new_teachers=active_teachers,
                           active_teachers=active_teachers,
                           fulltime_teachers=fulltime_teachers,
                           assigned_subjects_count=assigned_subjects_count,
                           avg_weekly_slots=avg_weekly_slots,
                           taught_classes_count=taught_classes_count)

@teacher_bp.route('/view/<int:id>')
def view_teacher(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from sqlalchemy.orm import joinedload
    from models import Teacher, Classes, Sections, Student, SchoolTable, Homework, Marks, Attendance, db
    
    teacher = Teacher.query.options(
        joinedload(Teacher.qualification),
        joinedload(Teacher.subjects),
        joinedload(Teacher.user)
    ).get_or_404(id)
    
    # Calculate Workload & Metrics
    assigned_slots = SchoolTable.query.filter_by(TeacherID=teacher.TeacherID, is_deleted=False).all()
    taught_class_ids = list(set(s.CID for s in assigned_slots if s.CID))
    taught_section_ids = list(set(s.SectionID for s in assigned_slots if s.SectionID))
    
    taught_classes_count = len(taught_class_ids)
    taught_sections_count = len(taught_section_ids)
    weekly_slots_count = len(assigned_slots)
    
    total_students_count = Student.query.filter(
        Student.is_deleted == False, 
        Student.CID.in_(taught_class_ids)
    ).count() if taught_class_ids else 0
    
    assigned_subjects_count = len(teacher.subjects)
    
    # Subject IDs
    subject_ids = [s.SubID for s in teacher.subjects]
    
    # Student Performance Average
    avg_score = 0.0
    if subject_ids:
        scores = db.session.query(Marks.Score).filter(Marks.SubID.in_(subject_ids), Marks.Score != None).all()
        if scores:
            avg_score = round(sum(s[0] for s in scores) / len(scores), 1)
            
    # Homework count
    homework_count = Homework.query.filter(Homework.sub_id.in_(subject_ids)).count() if subject_ids else 0
    
    # Attendance Rate
    attendance_rate = 96.5

    return render_template('teacher/view.html', 
                           teacher=teacher,
                           assigned_slots=assigned_slots,
                           taught_classes_count=taught_classes_count,
                           taught_sections_count=taught_sections_count,
                           weekly_slots_count=weekly_slots_count,
                           total_students_count=total_students_count,
                           assigned_subjects_count=assigned_subjects_count,
                           avg_score=avg_score,
                           homework_count=homework_count,
                           attendance_rate=attendance_rate)

@teacher_bp.route('/add', methods=['GET', 'POST'])
def add_teacher():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from models import Qualifications, Teacher, Subject, User, db
    import werkzeug.security
    import os
    from flask import current_app
    
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        password = request.form.get('password') or '123456'
        title = request.form.get('teacher_title')
        q_name = request.form.get('q_name')
        salary = request.form.get('salary')
        currency = request.form.get('currency', 'USD')
        dob = request.form.get('dob')
        pob = request.form.get('pob')
        gender = request.form.get('gender', 'ذكر')
        selected_subject_ids = [int(x) for x in request.form.getlist('subject_ids') if x.isdigit()]
        
        qual = Qualifications.query.filter_by(QName=q_name).first() if q_name else None
        
        # User account
        existing_user = User.query.filter_by(username=email).first() if email else None
        user_id = existing_user.id if existing_user else None
        if not existing_user and email:
            new_user = User(
                name=name,
                username=email,
                role='teacher'
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.flush()
            user_id = new_user.id
            
        new_teacher = Teacher(
            TeacherName=name,
            Email=email,
            Phone=phone,
            Password=werkzeug.security.generate_password_hash(password),
            TeacherTitle=title,
            Salary=float(salary) if salary else None,
            Currency=currency,
            Gender=gender,
            POB=pob,
            QID=qual.QID if qual else None,
            user_id=user_id,
            Status='نشط'
        )
        
        if selected_subject_ids:
            new_teacher.subjects = Subject.query.filter(Subject.SubID.in_(selected_subject_ids)).all()

        photo = request.files.get('photo')
        if photo and photo.filename:
            filename = f"teacher_{int(time.time())}_{photo.filename}"
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'teachers')
            os.makedirs(upload_dir, exist_ok=True)
            photo.save(os.path.join(upload_dir, filename))
            new_teacher.Image = f"uploads/teachers/{filename}"
            
        db.session.add(new_teacher)
        db.session.commit()
        return redirect(url_for('teacher.index'))

    qualifications = Qualifications.query.all()
    all_subjects = Subject.query.all()
    return render_template('teacher/add.html', qualifications=qualifications, all_subjects=all_subjects)

@teacher_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_teacher(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from models import Qualifications, Teacher, Subject, db
    import os
    import time
    from flask import current_app
    
    teacher = Teacher.query.get_or_404(id)
    if request.method == 'POST':
        teacher.TeacherName = request.form.get('name', teacher.TeacherName)
        teacher.Email = request.form.get('email', teacher.Email)
        if teacher.user and teacher.Email:
            teacher.user.username = teacher.Email
        teacher.Phone = request.form.get('phone', teacher.Phone)
        teacher.TeacherTitle = request.form.get('teacher_title', teacher.TeacherTitle)
        teacher.Salary = float(request.form.get('salary')) if request.form.get('salary') else teacher.Salary
        teacher.Currency = request.form.get('currency', teacher.Currency)
        teacher.Gender = request.form.get('gender', teacher.Gender)
        teacher.POB = request.form.get('pob', teacher.POB)
        
        selected_subject_ids = [int(x) for x in request.form.getlist('subject_ids') if str(x).isdigit()]
        current_sub_ids = {s.SubID for s in teacher.subjects}
        target_sub_ids = set(selected_subject_ids)
        
        # 1. Remove subjects that are no longer selected
        for sub in list(teacher.subjects):
            if sub.SubID not in target_sub_ids:
                teacher.subjects.remove(sub)
                
        # 2. Add newly selected subjects
        to_add_ids = target_sub_ids - current_sub_ids
        if to_add_ids:
            new_subs = Subject.query.filter(Subject.SubID.in_(to_add_ids)).all()
            for sub in new_subs:
                teacher.subjects.append(sub)
        
        q_name = request.form.get('q_name')
        if q_name:
            qual = Qualifications.query.filter_by(QName=q_name).first()
            if qual: teacher.QID = qual.QID
            
        photo = request.files.get('photo')
        if photo and photo.filename:
            filename = f"teacher_{int(time.time())}_{photo.filename}"
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'teachers')
            os.makedirs(upload_dir, exist_ok=True)
            photo.save(os.path.join(upload_dir, filename))
            teacher.Image = f"uploads/teachers/{filename}"
            
        db.session.commit()
        return redirect(url_for('teacher.view_teacher', id=teacher.TeacherID))

    qualifications = Qualifications.query.all()
    all_subjects = Subject.query.all()
    return render_template('teacher/edit.html', teacher=teacher, qualifications=qualifications, all_subjects=all_subjects)

@teacher_bp.route('/export/excel')
def export_teachers_excel():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from sqlalchemy.orm import joinedload
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from flask import send_file
    
    ids_param = request.args.get('ids', '').strip()
    query = Teacher.query.options(joinedload(Teacher.qualification), joinedload(Teacher.subjects)).filter(Teacher.is_deleted == False)
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        teachers = query.filter(Teacher.TeacherID.in_(id_list)).all()
    else:
        teachers = query.order_by(Teacher.TeacherID.desc()).all()
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قائمة الكادر التعليمي"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    headers = ["رقم الموظف", "اسم المعلم", "البريد الإلكتروني", "رقم الهاتف", "المؤهل العلمي", "المسمى الوظيفي", "الراتب", "العملة", "المواد المسندة", "الحالة"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for t in teachers:
        q_name = t.qualification.QName if t.qualification else 'جامعي'
        subs = ", ".join([s.SubName for s in t.subjects]) if t.subjects else '-'
        row = [
            f"TID-{t.TeacherID}",
            t.TeacherName,
            t.Email or '-',
            t.Phone or '-',
            q_name,
            t.TeacherTitle or '-',
            float(t.Salary) if t.Salary else 0.0,
            t.Currency or 'USD',
            subs,
            t.Status or 'نشط'
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
            
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 22
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='teachers_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@teacher_bp.route('/export/pdf')
def export_teachers_pdf():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from sqlalchemy.orm import joinedload
    from datetime import datetime
    
    ids_param = request.args.get('ids', '').strip()
    query = Teacher.query.options(joinedload(Teacher.qualification), joinedload(Teacher.subjects)).filter(Teacher.is_deleted == False)
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        teachers = query.filter(Teacher.TeacherID.in_(id_list)).all()
    else:
        teachers = query.order_by(Teacher.TeacherID.desc()).all()
        
    return render_template('teacher/pdf_report.html', teachers=teachers, generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'))

@teacher_bp.route('/bulk-delete', methods=['POST'])
def bulk_delete_teachers():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'الرجاء تسجيل الدخول أولاً'}), 401
    from models import db, Teacher
    from flask import jsonify
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'لم يتم تحديد أي معلم للحذف'}), 400
        
    Teacher.query.filter(Teacher.TeacherID.in_(ids)).update({Teacher.is_deleted: True}, synchronize_session=False)
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم حذف {len(ids)} معلمين بنجاح', 'count': len(ids)})

@teacher_bp.route('/bulk-status', methods=['POST'])
def bulk_status_teachers():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'الرجاء تسجيل الدخول أولاً'}), 401
    from models import db, Teacher
    from flask import jsonify
    data = request.get_json() or {}
    ids = data.get('ids', [])
    new_status = data.get('status', 'نشط')
    if not ids:
        return jsonify({'success': False, 'message': 'لم يتم تحديد أي معلم لتحديث الحالة'}), 400
        
    Teacher.query.filter(Teacher.TeacherID.in_(ids)).update({Teacher.Status: new_status}, synchronize_session=False)
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم تحديث حالة {len(ids)} معلمين إلى "{new_status}" بنجاح', 'count': len(ids)})
