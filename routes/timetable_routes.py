from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify
from flask_login import login_required, current_user
from models import db, SchoolTable, Classes, Sections, Subject, Teacher, Days, Lessons, Terms, Student, Attendance, Homework, ExamSchedule, Marks
from sqlalchemy.orm import joinedload
from datetime import datetime, timedelta

timetable_bp = Blueprint('timetable', __name__, url_prefix='/timetable')

@timetable_bp.route('/')
@login_required
def index():
    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        from services.teacher_timetable_service import (
            get_teacher_timetable_stats,
            get_teacher_today_schedule,
            get_teacher_weekly_schedule
        )

        stats = get_teacher_timetable_stats(current_user.id)
        today_schedule = get_teacher_today_schedule(current_user.id)
        weekly_schedule = get_teacher_weekly_schedule(current_user.id)

        return render_template('teacher/timetable.html',
                               stats=stats,
                               today_schedule=today_schedule,
                               weekly_schedule=weekly_schedule)

    # Admin Timetable View
    today = datetime.now().date()
    now_time_str = datetime.now().strftime('%H:%M')
    arabic_days = ['الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']
    today_day_name = arabic_days[today.weekday()]
    
    teacher = Teacher.query.options(joinedload(Teacher.subjects)).filter_by(user_id=current_user.id).first()
    
    req_teacher_id = request.args.get('teacher_id', type=int)
    target_teacher = None
    if req_teacher_id:
        target_teacher = Teacher.query.options(joinedload(Teacher.subjects)).filter_by(TeacherID=req_teacher_id, is_deleted=False).first()

    if not teacher and target_teacher:
        teacher = target_teacher
        
    all_teachers = Teacher.query.filter_by(is_deleted=False).order_by(Teacher.TeacherName.asc()).all()

    try:
        raw_class_id = request.args.get('class_id')
        class_id = int(raw_class_id) if raw_class_id and str(raw_class_id).isdigit() else None
    except (ValueError, TypeError):
        class_id = None

    subject_id = request.args.get('subject_id', type=int)
    
    selected_class = None
    class_not_found = False
    requested_class_id = raw_class_id if raw_class_id else None

    if class_id is not None:
        selected_class = Classes.query.filter_by(CID=class_id, is_deleted=False).first()
        if not selected_class:
            class_not_found = True
            class_id = None

    if not teacher:
        teacher_name = current_user.name
        teacher_title = 'إدارة النظام'
        teacher_status = 'نشط'
        subjects_str = 'كافة المواد الدراسية'
        words = teacher_name.split()
        initials = ". ".join([w[0] for w in words[:2]]) if len(words) >= 2 else 'م.أ'
        query = SchoolTable.query.options(
            joinedload(SchoolTable.subject),
            joinedload(SchoolTable.school_class),
            joinedload(SchoolTable.section),
            joinedload(SchoolTable.day),
            joinedload(SchoolTable.lesson)
        ).filter(SchoolTable.is_deleted == False)
        if class_id:
            query = query.filter(SchoolTable.CID == class_id)
        if subject_id:
            query = query.filter(SchoolTable.SubID == subject_id)
        slots = query.all()
        if not slots:
            from services.timetable_sync_service import sync_all_active_subject_timetable_slots
            sync_all_active_subject_timetable_slots()
            slots = query.all()
    else:
        teacher_id = teacher.TeacherID
        teacher_name = teacher.TeacherName
        teacher_title = teacher.TeacherTitle or 'معلم أكاديمي'
        teacher_status = teacher.Status or 'نشط'
        sub_names = [s.SubName for s in teacher.subjects] if teacher.subjects else []
        subjects_str = " | ".join(sub_names) if sub_names else 'المواد الدراسية'
        words = teacher_name.split()
        initials = ". ".join([w[0] for w in words[:2]]) if len(words) >= 2 else 'م.أ'
        
        query = SchoolTable.query.options(
            joinedload(SchoolTable.subject),
            joinedload(SchoolTable.school_class),
            joinedload(SchoolTable.section),
            joinedload(SchoolTable.day),
            joinedload(SchoolTable.lesson)
        ).filter_by(TeacherID=teacher_id, is_deleted=False)
        if class_id:
            query = query.filter(SchoolTable.CID == class_id)
        if subject_id:
            query = query.filter(SchoolTable.SubID == subject_id)
        slots = query.all()
        if not slots:
            from services.timetable_sync_service import sync_all_active_subject_timetable_slots
            sync_all_active_subject_timetable_slots()
            slots = query.all()

    teacher_info = {
        'TeacherName': teacher_name,
        'TeacherTitle': teacher_title,
        'Status': teacher_status,
        'subjects_str': subjects_str,
        'initials': initials,
        'school_name': 'مدرسة المستقبل الأهلية',
        'term_name': 'الفصل الدراسي الثاني - 2025/2026'
    }

    teacher_class_ids = list(set([s.CID for s in slots if s.CID]))
    teacher_section_ids = list(set([s.SectionID for s in slots if s.SectionID]))
    
    all_classes = Classes.query.filter_by(is_deleted=False).order_by(Classes.CID).all()

    if class_id:
        total_students = Student.query.filter(Student.CID == class_id, Student.is_deleted == False).count()
    else:
        total_students = Student.query.filter(Student.is_deleted == False).count()

    today_slots = [s for s in slots if s.day and s.day.DName == today_day_name]
    today_slots_sorted = sorted(
        today_slots, 
        key=lambda s: (s.lesson.StartTime if (s.lesson and s.lesson.StartTime) else '00:00')
    )
    
    today_lessons_count = len(today_slots_sorted)
    today_classes_count = len(set([s.CID for s in today_slots_sorted if s.CID]))
    
    current_lesson = {}
    next_lesson = {}
    current_slot_num = None
    current_slot_time = None
    next_slot_num = None
    next_slot_time = None
    daily_timeline = []
    finished_count = 0
    
    for idx, slot in enumerate(today_slots_sorted, start=1):
        start_t = slot.lesson.StartTime if (slot.lesson and slot.lesson.StartTime) else '08:00'
        end_t = slot.lesson.EndTime if (slot.lesson and slot.lesson.EndTime) else '08:45'
        sub_name = slot.subject.SubName if slot.subject else 'مادة تعليمية'
        cls_name = slot.school_class.CName if slot.school_class else ''
        sec_name = slot.section.SectionName if slot.section else ''
        full_cls = f"{cls_name} - {sec_name}".strip(" -")
        room_name = getattr(slot, 'RoomNo', None) or f"قاعة {200 + idx}"
        time_range = f"{start_t} - {end_t}"
        
        st_count = Student.query.filter(Student.CID == slot.CID, Student.is_deleted == False).count() if slot.CID else 0
        is_current = False
        is_next = False
        
        if end_t < now_time_str:
            status_text = 'انتهت'
            status_badge_class = 'secondary'
            finished_count += 1
        elif start_t <= now_time_str <= end_t:
            status_text = 'الحصة الحالية'
            status_badge_class = 'success'
            is_current = True
            current_slot_num = idx
            current_slot_time = time_range
            current_lesson = {
                'slot_num': idx,
                'subject': sub_name,
                'class': full_cls,
                'time': time_range,
                'room': room_name,
                'students_count': st_count,
                'attendance_rate': 0
            }
        else:
            if not next_lesson:
                status_text = 'الحصة القادمة'
                status_badge_class = 'warning'
                is_next = True
                next_slot_num = idx
                next_slot_time = time_range
                next_lesson = {
                    'slot_num': idx,
                    'subject': sub_name,
                    'class': full_cls,
                    'time': time_range,
                    'room': room_name,
                    'students_count': st_count
                }
            else:
                status_text = 'لاحقة'
                status_badge_class = 'info'
                
        daily_timeline.append({
            'num': idx,
            'subject_name': sub_name,
            'class_name': full_cls,
            'time': time_range,
            'room': room_name,
            'students_count': st_count,
            'status': status_text,
            'status_class': status_badge_class,
            'is_current': is_current,
            'is_next': is_next
        })

    if not current_lesson and today_slots_sorted:
        first_s = today_slots_sorted[0]
        start_t = first_s.lesson.StartTime if (first_s.lesson and first_s.lesson.StartTime) else '08:55'
        end_t = first_s.lesson.EndTime if (first_s.lesson and first_s.lesson.EndTime) else '09:40'
        sub_name = first_s.subject.SubName if first_s.subject else 'مادة تعليمية'
        cls_name = first_s.school_class.CName if first_s.school_class else ''
        sec_name = first_s.section.SectionName if first_s.section else ''
        current_lesson = {
            'slot_num': 1,
            'subject': sub_name,
            'class': f"{cls_name} - {sec_name}".strip(" -"),
            'time': f"{start_t} - {end_t}",
            'room': getattr(first_s, 'RoomNo', None) or 'قاعة 201',
            'students_count': total_students or 0,
            'attendance_rate': 0
        }
        current_slot_num = 1
        current_slot_time = f"{start_t} - {end_t}"

    occupancy_rate = min(100, round((today_lessons_count / 6.0) * 100)) if today_lessons_count > 0 else 0

    week_days = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
    lessons_list = Lessons.query.filter_by(is_deleted=False).order_by(Lessons.LessonID.asc()).all()
    if not lessons_list:
        _ensure_default_days_and_lessons()
        lessons_list = Lessons.query.filter_by(is_deleted=False).order_by(Lessons.LessonID.asc()).all()

    weekly_matrix = {}
    for day in week_days:
        weekly_matrix[day] = {}

    for s in slots:
        if s.day and s.day.DName in weekly_matrix and s.lesson:
            sub_name = s.subject.SubName if s.subject else ''
            cls_name = s.school_class.CName if s.school_class else ''
            sec_name = s.section.SectionName if s.section else ''
            full_cls = f"{cls_name} - {sec_name}".strip(" -")
            start_t = s.lesson.StartTime if s.lesson.StartTime else '08:00'
            lesson_name = s.lesson.LessonName or f"الحصة {s.LessonID}"
            weekly_matrix[s.day.DName][s.LessonID] = {
                'subject': sub_name,
                'class': full_cls,
                'time': start_t,
                'lesson_name': lesson_name,
                'room': getattr(s, 'RoomNo', None) or 'قاعة 201'
            }

    if class_id:
        today_present = Attendance.query.join(Student, Attendance.SID == Student.SID).filter(Attendance.Date == today, Student.CID == class_id, Attendance.Status.in_(['حاضر', 'متأخر'])).count()
        today_absent = Attendance.query.join(Student, Attendance.SID == Student.SID).filter(Attendance.Date == today, Student.CID == class_id, Attendance.Status == 'غائب').count()
    else:
        today_present = Attendance.query.filter(Attendance.Date == today, Attendance.Status.in_(['حاضر', 'متأخر'])).count()
        today_absent = Attendance.query.filter(Attendance.Date == today, Attendance.Status == 'غائب').count()
        
    total_att = today_present + today_absent
    att_rate = round((today_present / total_att * 100), 1) if total_att > 0 else 0.0
    
    kpi_data = {
        'today_lessons_count': today_lessons_count,
        'current_slot_num': current_slot_num or ('1' if today_slots_sorted else '—'),
        'current_slot_time': current_slot_time or ('—' if not today_slots_sorted else ''),
        'next_slot_num': next_slot_num or ('—' if len(today_slots_sorted) <= 1 else ''),
        'next_slot_time': next_slot_time or '—',
        'total_students': total_students,
        'occupancy_rate': occupancy_rate or 0,
        'taught_classes_count': today_classes_count or len(teacher_class_ids)
    }
    
    today_summary = {
        'completed_str': f"{finished_count} من {today_lessons_count}",
        'attendance_rate': att_rate,
        'absent_count': today_absent or 0,
        'present_count': today_present or 0
    }

    return render_template('timetable/index.html',
                           teacher_info=teacher_info,
                           kpi=kpi_data,
                           current_lesson=current_lesson,
                           next_lesson=next_lesson,
                           daily_timeline=daily_timeline,
                           weekly_matrix=weekly_matrix,
                           week_days=week_days,
                           lessons_list=lessons_list,
                           today_summary=today_summary,
                           today_date=today.strftime('%Y-%m-%d'),
                           today_day_name=today_day_name,
                           all_classes=all_classes,
                           selected_class=selected_class,
                           selected_class_id=class_id,
                           class_not_found=class_not_found,
                           requested_class_id=requested_class_id)

@timetable_bp.route('/api/drawer/<int:slot_id>')
@login_required
def lesson_drawer_api(slot_id):
    from services.teacher_timetable_service import get_lesson_drawer_data
    data = get_lesson_drawer_data(slot_id, current_user.id)
    if not data:
        return jsonify({'error': 'Lesson not found or access forbidden'}), 403
    return jsonify(data)

@timetable_bp.route('/export/excel')
@login_required
def export_timetable_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from flask import send_file

    class_id = request.args.get('class_id', type=int)
    selected_class = Classes.query.filter_by(CID=class_id, is_deleted=False).first() if class_id else None

    query = SchoolTable.query.options(
        joinedload(SchoolTable.subject),
        joinedload(SchoolTable.school_class),
        joinedload(SchoolTable.section),
        joinedload(SchoolTable.day),
        joinedload(SchoolTable.lesson)
    ).filter(SchoolTable.is_deleted == False)

    if class_id:
        query = query.filter(SchoolTable.CID == class_id)

    slots = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الجدول الدراسي"
    ws.views.sheetView[0].rightToLeft = True

    # Title Banner
    title_text = f"الجدول الدراسي الأسبوعي - {selected_class.CName}" if selected_class else "الجدول الدراسي الشامل للمدرسة"
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = Font(name='Arial', size=15, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = ['اليوم', 'الحصة / الوقت', 'المادة الدراسية', 'الصف الدراسي', 'الشعبة', 'القاعة / المكان']
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for slot in slots:
        day_name = slot.day.DName if slot.day else 'غير محدد'
        lesson_str = f"{slot.lesson.LessonName} ({slot.lesson.StartTime} - {slot.lesson.EndTime})" if slot.lesson else 'غير محدد'
        sub_name = slot.subject.SubName if slot.subject else 'غير محدد'
        cls_name = slot.school_class.CName if slot.school_class else 'غير محدد'
        sec_name = slot.section.SectionName if slot.section else 'جميع الشعب'
        room_name = getattr(slot, 'RoomNo', None) or 'قاعة دراسية'

        ws.append([day_name, lesson_str, sub_name, cls_name, sec_name, room_name])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"timetable_class_{class_id}.xlsx" if class_id else "timetable_full.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _ensure_default_days_and_lessons():
    if Days.query.count() == 0:
        day_names = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
        for d in day_names:
            db.session.add(Days(DName=d))
        db.session.commit()

    if Lessons.query.count() == 0:
        lesson_times = [
            ('الحصة الأولى', '08:00', '08:45'),
            ('الحصة الثانية', '08:45', '09:30'),
            ('الحصة الثالثة', '09:30', '10:15'),
            ('الحصة الرابعة', '10:30', '11:15'),
            ('الحصة الخامسة', '11:15', '12:00'),
            ('الحصة السادسة', '12:00', '12:45'),
            ('الحصة السابعة', '12:45', '01:30'),
        ]
        for name, start, end in lesson_times:
            db.session.add(Lessons(LessonName=name, StartTime=start, EndTime=end))
        db.session.commit()

@timetable_bp.route('/builder', methods=['GET'])
@login_required
def builder():
    _ensure_default_days_and_lessons()

    classes = Classes.query.filter_by(is_deleted=False).order_by(Classes.CName.asc()).all()
    sections = Sections.query.filter_by(is_deleted=False).all()
    teachers = Teacher.query.filter_by(is_deleted=False).order_by(Teacher.TeacherName.asc()).all()
    subjects = Subject.query.filter_by(is_deleted=False).order_by(Subject.SubName.asc()).all()
    terms = Terms.query.filter_by(is_deleted=False).all()

    days = Days.query.filter_by(is_deleted=False).order_by(Days.DayID.asc()).all()
    lessons = Lessons.query.filter_by(is_deleted=False).order_by(Lessons.LessonID.asc()).all()

    sel_class_id = request.args.get('class_id', type=int) or (classes[0].CID if classes else None)
    sel_sec_id = request.args.get('section_id', type=int) or (sections[0].SectionID if sections else None)
    sel_term_id = request.args.get('term_id', type=int) or (terms[0].T_ID if terms else None)

    slots_query = SchoolTable.query.options(
        joinedload(SchoolTable.subject),
        joinedload(SchoolTable.teacher),
        joinedload(SchoolTable.school_class),
        joinedload(SchoolTable.section),
        joinedload(SchoolTable.day),
        joinedload(SchoolTable.lesson)
    ).filter(SchoolTable.is_deleted == False)

    if sel_class_id:
        slots_query = slots_query.filter(SchoolTable.CID == sel_class_id)
    if sel_sec_id:
        slots_query = slots_query.filter(SchoolTable.SectionID == sel_sec_id)

    raw_slots = slots_query.all()

    grid = {}
    for slot in raw_slots:
        key = f"{slot.DayID}_{slot.LessonID}"
        grid[key] = {
            'slot_id': slot.SchoolTableID,
            'subject_id': slot.SubID,
            'subject_name': slot.subject.SubName if slot.subject else '',
            'color': getattr(slot.subject, 'Color', '#3b82f6') if slot.subject else '#3b82f6',
            'teacher_id': slot.TeacherID,
            'teacher_name': slot.teacher.TeacherName if slot.teacher else '',
            'class_name': slot.school_class.CName if slot.school_class else '',
            'section_name': slot.section.SectionName if slot.section else ''
        }

    return render_template(
        'timetable/builder.html',
        classes=classes,
        sections=sections,
        teachers=teachers,
        subjects=subjects,
        terms=terms,
        days=days,
        lessons=lessons,
        sel_class_id=sel_class_id,
        sel_sec_id=sel_sec_id,
        sel_term_id=sel_term_id,
        grid=grid
    )

@timetable_bp.route('/api/check-conflict', methods=['POST'])
@login_required
def check_conflict():
    data = request.get_json(silent=True) or {}
    teacher_id = data.get('teacher_id')
    day_id = data.get('day_id')
    lesson_id = data.get('lesson_id')
    class_id = data.get('class_id')
    section_id = data.get('section_id')
    current_slot_id = data.get('slot_id')

    if not teacher_id or not day_id or not lesson_id:
        return jsonify({'has_conflict': False})

    # Check Teacher Conflict
    t_query = SchoolTable.query.filter(
        SchoolTable.TeacherID == teacher_id,
        SchoolTable.DayID == day_id,
        SchoolTable.LessonID == lesson_id,
        SchoolTable.is_deleted == False
    )
    if current_slot_id:
        t_query = t_query.filter(SchoolTable.SchoolTableID != current_slot_id)

    conflict_t = t_query.first()
    if conflict_t:
        t_name = conflict_t.teacher.TeacherName if conflict_t.teacher else 'المعلم'
        c_name = conflict_t.school_class.CName if conflict_t.school_class else ''
        s_name = conflict_t.section.SectionName if conflict_t.section else ''
        return jsonify({
            'has_conflict': True,
            'type': 'teacher',
            'message': f"⚠️ المعلم ({t_name}) معين في حصة أخرى بهذا التوقيت في ({c_name} - {s_name})!"
        })

    # Check Class / Section Conflict
    if class_id and section_id:
        c_query = SchoolTable.query.filter(
            SchoolTable.CID == class_id,
            SchoolTable.SectionID == section_id,
            SchoolTable.DayID == day_id,
            SchoolTable.LessonID == lesson_id,
            SchoolTable.is_deleted == False
        )
        if current_slot_id:
            c_query = c_query.filter(SchoolTable.SchoolTableID != current_slot_id)

        conflict_c = c_query.first()
        if conflict_c:
            sub_name = conflict_c.subject.SubName if conflict_c.subject else ''
            return jsonify({
                'has_conflict': True,
                'type': 'class',
                'message': f"⚠️ هذا الصف والشعبة لديه حصة مجدولة بالفعل ({sub_name}) في هذا التوقيت!"
            })

    return jsonify({'has_conflict': False})

@timetable_bp.route('/api/assign-slot', methods=['POST'])
@login_required
def assign_slot():
    data = request.get_json(silent=True) or {}
    class_id = data.get('class_id')
    section_id = data.get('section_id')
    day_id = data.get('day_id')
    lesson_id = data.get('lesson_id')
    subject_id = data.get('subject_id')
    teacher_id = data.get('teacher_id')
    term_id = data.get('term_id')
    slot_id = data.get('slot_id')

    if not all([class_id, section_id, day_id, lesson_id, subject_id, teacher_id]):
        return jsonify({'error': 'جميع الحقول مطلوبة لتسكين الحصة في الجدول'}), 400

    existing_slot = SchoolTable.query.filter_by(
        CID=class_id, SectionID=section_id, DayID=day_id, LessonID=lesson_id, is_deleted=False
    ).first()
    target_slot_id = slot_id or (existing_slot.SchoolTableID if existing_slot else None)

    # Conflict Check
    t_conflict = SchoolTable.query.filter(
        SchoolTable.TeacherID == teacher_id,
        SchoolTable.DayID == day_id,
        SchoolTable.LessonID == lesson_id,
        SchoolTable.is_deleted == False
    )
    if target_slot_id:
        t_conflict = t_conflict.filter(SchoolTable.SchoolTableID != target_slot_id)

    if t_conflict.first():
        conflict_obj = t_conflict.first()
        t_name = conflict_obj.teacher.TeacherName if conflict_obj.teacher else ''
        c_name = conflict_obj.school_class.CName if conflict_obj.school_class else ''
        return jsonify({'error': f"عفواً، المعلم ({t_name}) لديه حصة أخرى في نفس التوقيت في ({c_name})"}), 400

    if slot_id:
        slot = db.session.get(SchoolTable, slot_id)
        if not slot:
            return jsonify({'error': 'الحصة غير موجودة'}), 404
        slot.CID = class_id
        slot.SectionID = section_id
        slot.DayID = day_id
        slot.LessonID = lesson_id
        slot.SubID = subject_id
        slot.TeacherID = teacher_id
        if term_id:
            slot.T_ID = term_id
    elif existing_slot:
        slot = existing_slot
        slot.SubID = subject_id
        slot.TeacherID = teacher_id
        if term_id:
            slot.T_ID = term_id
    else:
        slot = SchoolTable(
            CID=class_id,
            SectionID=section_id,
            DayID=day_id,
            LessonID=lesson_id,
            SubID=subject_id,
            TeacherID=teacher_id,
            T_ID=term_id
        )
        db.session.add(slot)

    db.session.commit()

    # Return refreshed slot object
    slot_ref = db.session.get(SchoolTable, slot.SchoolTableID)
    return jsonify({
        'success': True,
        'message': 'تم تسكين الحصة في الجدول بنجاح 🟢',
        'slot': {
            'slot_id': slot_ref.SchoolTableID,
            'subject_name': slot_ref.subject.SubName if slot_ref.subject else '',
            'teacher_name': slot_ref.teacher.TeacherName if slot_ref.teacher else '',
            'color': getattr(slot_ref.subject, 'Color', '#3b82f6') if slot_ref.subject else '#3b82f6'
        }
    })

@timetable_bp.route('/api/delete-slot/<int:slot_id>', methods=['POST', 'DELETE'])
@login_required
def delete_slot(slot_id):
    slot = db.session.get(SchoolTable, slot_id)
    if not slot:
        return jsonify({'error': 'الحصة غير موجودة'}), 404

    slot.is_deleted = True
    db.session.commit()
    return jsonify({'success': True, 'message': 'تم إزالة الحصة من الجدول بنجاح'})
