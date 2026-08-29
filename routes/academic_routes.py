from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from sqlalchemy.exc import IntegrityError
from models import db, Classes, Sections, Subject, Days, Lessons, Terms, TypeExams, Student, ExamSchedule
from models.timetable import SchoolTable
from utils.decorators import admin_required

academic_bp = Blueprint('academic', __name__, url_prefix='/academic')

def _get_class_teachers_count(cid, section_id=None):
    from models import Teacher
    query = db.session.query(SchoolTable.TeacherID).filter(
        SchoolTable.CID == cid,
        SchoolTable.is_deleted == False,
        SchoolTable.TeacherID.isnot(None)
    )
    if section_id:
        query = query.filter(SchoolTable.SectionID == section_id)
    t_ids = [t[0] for t in query.distinct().all() if t[0]]
    if not t_ids:
        return 0
    return Teacher.query.filter(Teacher.TeacherID.in_(t_ids), Teacher.is_deleted == False).count()

@academic_bp.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    # Redirect to classes by default
    return redirect(url_for('academic.classes'))

@academic_bp.route('/classes')
def classes():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    
    classes_list = Classes.query.filter_by(is_deleted=False).order_by(Classes.CID.asc()).all()
    active_class_ids = [c.CID for c in classes_list]
    
    from models.academic import ClassesSections
    if active_class_ids:
        sections_list = db.session.query(Sections).join(
            ClassesSections, Sections.SectionID == ClassesSections.c.SectionID
        ).filter(
            ClassesSections.c.CID.in_(active_class_ids),
            Sections.is_deleted == False
        ).distinct().all()
    else:
        sections_list = []
    
    total_classes = len(classes_list)
    total_sections = len(sections_list)
    active_classes = sum(1 for c in classes_list if getattr(c, 'Status', 'نشط') in ['نشط', None])
    active_sections = sum(1 for s in sections_list if getattr(s, 'Status', 'نشط') in ['نشط', None])
    
    total_students = Student.query.filter_by(is_deleted=False).count()
    avg_students_per_class = round(total_students / total_classes, 1) if total_classes > 0 else 0
    
    # Calculate overall occupancy strictly for classes where MaxStudents is defined
    total_capacity = sum(c.MaxStudents for c in classes_list if c.MaxStudents and c.MaxStudents > 0)
    overall_occupancy_rate = round((total_students / total_capacity) * 100, 1) if total_capacity > 0 else None
    
    # Enrich each class object with real dynamic metrics
    for c in classes_list:
        c.linked_sections = [s for s in c.sections if not getattr(s, 'is_deleted', False)]
        c.students_count = Student.query.filter_by(CID=c.CID, is_deleted=False).count()
        c.teachers_count = _get_class_teachers_count(c.CID)
        c.subjects_count = len([sub for sub in c.subjects if not getattr(sub, 'is_deleted', False)])
        
        # Calculate occupancy rate (MUST be None if MaxStudents is None or 0)
        if c.MaxStudents and c.MaxStudents > 0:
            c.occupancy_percentage = round((c.students_count / c.MaxStudents) * 100, 1)
            if c.occupancy_percentage < 70:
                c.occupancy_color = 'success'
            elif c.occupancy_percentage <= 90:
                c.occupancy_color = 'warning'
            else:
                c.occupancy_color = 'danger'
        else:
            c.occupancy_percentage = None
            c.occupancy_color = 'secondary'
            
    middle_school = sum(1 for c in classes_list if c.Stage == 'المتوسطة')
    high_school = sum(1 for c in classes_list if c.Stage == 'الثانوية')
    primary_school = sum(1 for c in classes_list if c.Stage == 'الأساسية')
    
    return render_template('academic/classes.html', 
                           classes=classes_list, 
                           sections=sections_list,
                           middle_school=middle_school,
                           high_school=high_school,
                           primary_school=primary_school,
                           total_sections=total_sections,
                           total_classes=total_classes,
                           active_classes=active_classes,
                           active_sections=active_sections,
                           avg_students_per_class=avg_students_per_class,
                           overall_occupancy_rate=overall_occupancy_rate)

@academic_bp.route('/subjects')
def subjects():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    
    from models.teacher import Teacher
    class_id = request.args.get('class_id', type=int)
    classes = Classes.query.filter_by(is_deleted=False).order_by(Classes.CID.asc()).all()
    for c in classes:
        c.linked_sections = [s for s in c.sections if not getattr(s, 'is_deleted', False)]
        c.students_count = Student.query.filter_by(CID=c.CID, is_deleted=False).count()
        
    teachers_list = Teacher.query.filter_by(is_deleted=False).all()
    
    query = Subject.query
    if hasattr(Subject, 'is_deleted'):
        query = query.filter_by(is_deleted=False)
        
    if class_id:
        from models.academic import ClassSubject
        subject_ids = [cs[0] for cs in db.session.query(ClassSubject.c.SubID).filter(ClassSubject.c.CID == class_id).all()]
        subjects_list = query.filter(Subject.SubID.in_(subject_ids)).order_by(Subject.SubID.asc()).all() if subject_ids else []
    else:
        subjects_list = query.order_by(Subject.SubID.asc()).all()
        
    total_subjects = len(subjects_list)
    active_subjects = sum(1 for s in subjects_list if getattr(s, 'Status', 'نشط') in ['نشط', None])
    inactive_subjects = sum(1 for s in subjects_list if getattr(s, 'Status', 'نشط') == 'غير نشط')
    
    from models import SchoolTable
    from models.academic import ClassSubject, TeacherSubject
    
    total_classes = len(classes)
    linked_classes_count = db.session.query(ClassSubject.c.CID)\
        .join(Classes, ClassSubject.c.CID == Classes.CID)\
        .join(Subject, ClassSubject.c.SubID == Subject.SubID)\
        .filter(Classes.is_deleted == False, Subject.is_deleted == False).distinct().count()
        
    assigned_teachers_count = db.session.query(TeacherSubject.c.TeacherID)\
        .join(Teacher, TeacherSubject.c.TeacherID == Teacher.TeacherID)\
        .join(Subject, TeacherSubject.c.SubID == Subject.SubID)\
        .filter(Teacher.is_deleted == False, Subject.is_deleted == False).distinct().count()
    
    total_valid_links = db.session.query(ClassSubject)\
        .join(Classes, ClassSubject.c.CID == Classes.CID)\
        .join(Subject, ClassSubject.c.SubID == Subject.SubID)\
        .filter(Classes.is_deleted == False, Subject.is_deleted == False).count()
        
    avg_subjects_per_class = round(total_valid_links / total_classes, 1) if total_classes > 0 else 0.0
    
    from models import SchoolTable, Terms, ExamSchedule, Homework
    from models.grade import Marks, HomeworkMarks
    from models.academic import ClassSubject, TeacherSubject
    from sqlalchemy.orm import joinedload

    current_term = Terms.query.filter_by(is_deleted=False).first()
    academic_year = current_term.AcademicYear if (current_term and current_term.AcademicYear) else '2025-2026'

    for s in subjects_list:
        # Linked Classes
        c_ids_cs = [c[0] for c in db.session.query(ClassSubject.c.CID).filter(ClassSubject.c.SubID == s.SubID).all() if c[0]]
        c_ids_st = [st[0] for st in db.session.query(SchoolTable.CID).filter(SchoolTable.SubID == s.SubID, SchoolTable.is_deleted == False).all() if st[0]]
        all_class_ids = list(set(c_ids_cs + c_ids_st))

        linked_cls_objs = Classes.query.filter(Classes.CID.in_(all_class_ids), Classes.is_deleted == False).all() if all_class_ids else []
        for cls in linked_cls_objs:
            cls.sections_count = len([sec for sec in cls.sections if not getattr(sec, 'is_deleted', False)])
            cls.students_count = Student.query.filter_by(CID=cls.CID, is_deleted=False).count()
            cls.max_students = cls.MaxStudents or 40
            cls.occupancy_percentage = round((cls.students_count / cls.max_students) * 100.0, 1) if cls.max_students > 0 else 0.0

        s.linked_classes = linked_cls_objs
        s.linked_classes_count = len(linked_cls_objs)

        # Assigned Teachers
        t_ids_ts = [t[0] for t in db.session.query(TeacherSubject.c.TeacherID)\
            .join(Teacher, TeacherSubject.c.TeacherID == Teacher.TeacherID)\
            .filter(TeacherSubject.c.SubID == s.SubID, Teacher.is_deleted == False).distinct().all() if t[0]]
            
        t_ids_st = [t[0] for t in db.session.query(SchoolTable.TeacherID)\
            .join(Teacher, SchoolTable.TeacherID == Teacher.TeacherID)\
            .filter(SchoolTable.SubID == s.SubID, SchoolTable.is_deleted == False, Teacher.is_deleted == False).distinct().all() if t[0]]
            
        all_teacher_ids = list(set(t_ids_ts + t_ids_st))
        assigned_teachers = Teacher.query.filter(Teacher.TeacherID.in_(all_teacher_ids), Teacher.is_deleted == False).all() if all_teacher_ids else []
        s.assigned_teachers = assigned_teachers
        s.assigned_teachers_count = len(assigned_teachers)

        # Students count
        if all_class_ids:
            s.students_count = Student.query.filter(Student.CID.in_(all_class_ids), Student.is_deleted == False).count()
        else:
            s.students_count = 0

        # Weekly Timetable Slots
        st_slots = SchoolTable.query.options(
            joinedload(SchoolTable.day),
            joinedload(SchoolTable.lesson),
            joinedload(SchoolTable.school_class),
            joinedload(SchoolTable.section),
            joinedload(SchoolTable.teacher)
        ).filter(SchoolTable.SubID == s.SubID, SchoolTable.is_deleted == False).all()

        s.weekly_slots_count = len(st_slots)
        
        timetable_slots_payload = []
        for slot in st_slots:
            timetable_slots_payload.append({
                'day': slot.day.DName if slot.day else 'الأحد',
                'lesson': slot.lesson.LessonName if slot.lesson else 'الحصة 1',
                'className': slot.school_class.CName if slot.school_class else 'الصف',
                'sectionName': slot.section.SectionName if slot.section else '',
                'teacherName': slot.teacher.TeacherName if slot.teacher else 'معلم'
            })
        s.timetable_slots_payload = timetable_slots_payload

        # Average score & Pass rate
        ex_marks = Marks.query.filter(Marks.SubID == s.SubID, Marks.is_deleted == False, Marks.Score.isnot(None)).all()
        hw_marks = HomeworkMarks.query.filter(HomeworkMarks.SubID == s.SubID, HomeworkMarks.is_deleted == False, HomeworkMarks.Score.isnot(None)).all()
        
        all_pcts = []
        for m in ex_marks:
            max_s = float(m.MaxScore) if m.MaxScore else 100.0
            all_pcts.append((float(m.Score) / max_s * 100.0) if max_s > 0 else float(m.Score))
        for h in hw_marks:
            if h.Percentage is not None:
                all_pcts.append(float(h.Percentage))
            else:
                sc = float(h.Score)
                max_s = float(h.MaxScore) if h.MaxScore else (10.0 if sc <= 10.0 else 100.0)
                all_pcts.append((sc / max_s * 100.0) if max_s > 0 else (sc * 10.0 if sc <= 10.0 else sc))
                
        if all_pcts:
            s.avg_score = round(sum(all_pcts) / len(all_pcts), 1)
            passed = sum(1 for p in all_pcts if p >= 60.0)
            s.pass_rate = round((passed / len(all_pcts)) * 100.0, 1)
        else:
            s.avg_score = None
            s.pass_rate = None

        # Activity Timeline
        timeline = []
        recent_hw = Homework.query.filter_by(sub_id=s.SubID).order_by(Homework.id.desc()).first()
        if recent_hw:
            timeline.append({
                'title': f'تم إضافة واجب دراسي: {recent_hw.title}',
                'time': recent_hw.created_at.strftime('%Y-%m-%d') if hasattr(recent_hw, 'created_at') and recent_hw.created_at else 'مؤخراً',
                'icon': 'fa-book-open',
                'color': 'bg-primary text-white'
            })
        recent_ex = ExamSchedule.query.filter_by(SubID=s.SubID).order_by(ExamSchedule.ScheduleID.desc()).first()
        if recent_ex:
            timeline.append({
                'title': f'تم جدولة اختبار: {recent_ex.ExamName}',
                'time': recent_ex.ExamDate.strftime('%Y-%m-%d') if recent_ex.ExamDate else 'مؤخراً',
                'icon': 'fa-file-signature',
                'color': 'bg-danger text-white'
            })
        if st_slots:
            timeline.append({
                'title': f'تأكيد {len(st_slots)} حصة دراسية مجدولة بالجدول الأسبوعي',
                'time': 'الجدول المعتمد',
                'icon': 'fa-calendar-check',
                'color': 'bg-success text-white'
            })
        if not timeline:
            timeline.append({
                'title': 'تم اعتماد المادة الدراسية بالخطة الأكاديمية',
                'time': 'السجل الرئيسي',
                'icon': 'fa-check-circle',
                'color': 'bg-info text-white'
            })
        s.activity_timeline = timeline
        s.academic_year = academic_year
        
    selected_class = Classes.query.filter_by(CID=class_id).first() if class_id else None
    
    return render_template('academic/subjects.html', 
                           subjects=subjects_list,
                           classes=classes,
                           teachers=teachers_list,
                           selected_class_id=class_id,
                           selected_class=selected_class,
                           total_subjects=total_subjects,
                           active_subjects=active_subjects,
                           inactive_subjects=inactive_subjects,
                           linked_classes_count=linked_classes_count,
                           assigned_teachers_count=assigned_teachers_count,
                           avg_subjects_per_class=avg_subjects_per_class)

@academic_bp.route('/add_class', methods=['POST'])
def add_class():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    name = request.form.get('name', '').strip()
    stage = request.form.get('stage', '').strip()
    if not name:
        flash('يرجى كتابة اسم الصف الدراسي', 'warning')
        return redirect(url_for('academic.classes'))
        
    query = Classes.query.filter_by(CName=name)
    if stage:
        query = query.filter_by(Stage=stage)
    existing_class = query.first()

    if existing_class:
        if getattr(existing_class, 'is_deleted', False):
            existing_class.is_deleted = False
            if stage:
                existing_class.Stage = stage
            try:
                db.session.commit()
                flash(f'تمت استعادة وتفعيل الصف "{name}" بنجاح', 'success')
            except Exception:
                db.session.rollback()
                flash('حدث خطأ أثناء تفعيل الصف', 'danger')
        else:
            stage_desc = f' ({stage})' if stage else ''
            flash(f'الصف "{name}"{stage_desc} موجود بالفعل في النظام', 'warning')
        return redirect(url_for('academic.classes'))
        
    new_class = Classes(CName=name, Stage=stage)
    try:
        db.session.add(new_class)
        db.session.commit()
        flash(f'تمت إضافة الصف "{name}" بنجاح', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'الصف "{name}" موجود بالفعل أو حدث تعارض في البيانات', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة الصف: {str(e)}', 'danger')
        
    return redirect(url_for('academic.classes'))

@academic_bp.route('/add_section', methods=['POST'])
def add_section():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    name = request.form.get('name', '').strip()
    class_id = request.form.get('class_id')
    if not name or not class_id:
        flash('يرجى إدخال اسم الشعبة واختيار الصف الدراسي', 'warning')
        return redirect(url_for('academic.classes'))
        
    c = Classes.query.get(class_id)
    if not c:
        flash('الصف الدراسي غير موجود', 'danger')
        return redirect(url_for('academic.classes'))
        
    try:
        new_section = Sections(SectionName=name)
        new_section.classes.append(c)
        db.session.add(new_section)
        db.session.commit()
        flash(f'تمت إضافة الشعبة "{name}" بنجاح للصف "{c.CName}"', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'الشعبة "{name}" موجودة بالفعل أو تعذر إضافتها', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة الشعبة: {str(e)}', 'danger')
        
    return redirect(url_for('academic.classes'))

@academic_bp.route('/class/<int:id>/sections', methods=['GET'])
def get_class_sections_api(id):
    from flask import jsonify
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from models import SchoolTable, ClassesSections

    try:
        class_id = int(id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'رقم الصف الدراسي غير صالح'}), 400

    c = Classes.query.filter_by(CID=class_id, is_deleted=False).first()
    if not c:
        return jsonify({'success': False, 'message': f'الصف الدراسي رقم ({id}) غير موجود في النظام'}), 404
    
    raw_sections = db.session.query(Sections).join(ClassesSections, Sections.SectionID == ClassesSections.c.SectionID)\
        .filter(ClassesSections.c.CID == class_id).all()
    sections_list = []
    
    for sec in raw_sections:
        if not getattr(sec, 'is_deleted', False):
            st_count = Student.query.filter(Student.SectionID == sec.SectionID, Student.CID == c.CID, Student.is_deleted == False).count()
            tb_count = SchoolTable.query.filter(SchoolTable.SectionID == sec.SectionID, SchoolTable.CID == c.CID, SchoolTable.is_deleted == False).count()
            t_count = _get_class_teachers_count(c.CID, sec.SectionID)

            sections_list.append({
                'id': sec.SectionID,
                'name': sec.SectionName,
                'maxStudents': sec.MaxStudents or 40,
                'studentsCount': st_count,
                'teachersCount': t_count,
                'timetableCount': tb_count
            })
            
    return jsonify({
        'success': True,
        'class': {
            'id': c.CID,
            'name': c.CName,
            'stage': c.Stage or 'المرحلة العامة'
        },
        'sections': sections_list
    })

@academic_bp.route('/class/<int:id>/sections/add', methods=['POST'])
def add_class_section_api(id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    from models.academic import ClassesSections
    c = Classes.query.get_or_404(id)
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': 'يرجى كتابة اسم الشعبة'}), 400
        
    existing_sections = db.session.query(Sections).join(ClassesSections, Sections.SectionID == ClassesSections.c.SectionID)\
        .filter(ClassesSections.c.CID == id).all()
    for sec in existing_sections:
        if sec.SectionName and sec.SectionName.strip().lower() == name.lower():
            if not getattr(sec, 'is_deleted', False):
                return jsonify({'success': False, 'message': f'الشعبة "{name}" موجودة بالفعل لصف {c.CName}'}), 400
            
    try:
        new_sec = Sections(SectionName=name)
        db.session.add(new_sec)
        db.session.flush()
        db.session.execute(ClassesSections.insert().values(CID=c.CID, SectionID=new_sec.SectionID))
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'تمت إضافة الشعبة "{name}" بنجاح للصف "{c.CName}"',
            'section': {
                'id': new_sec.SectionID,
                'name': new_sec.SectionName,
                'studentsCount': 0,
                'timetableCount': 0
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ أثناء إضافة الشعبة: {str(e)}'}), 500

@academic_bp.route('/section/<int:sec_id>/edit', methods=['POST'])
def edit_section_api(sec_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    sec = Sections.query.get_or_404(sec_id)
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': 'اسم الشعبة لا يمكن أن يكون فارغاً'}), 400
        
    sec.SectionName = name
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم تحديث اسم الشعبة بنجاح إلى "{name}"'})

@academic_bp.route('/class/<int:class_id>/section/<int:sec_id>/delete', methods=['POST'])
def delete_section_api(class_id, sec_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    from models import SchoolTable, Homework, ExamSchedule
    c = Classes.query.get_or_404(class_id)
    sec = Sections.query.get_or_404(sec_id)
    
    students_count = Student.query.filter(Student.SectionID == sec.SectionID, Student.CID == c.CID, Student.is_deleted == False).count()
    timetable_count = SchoolTable.query.filter(SchoolTable.SectionID == sec.SectionID, SchoolTable.CID == c.CID, SchoolTable.is_deleted == False).count()
    homework_count = Homework.query.filter(Homework.section_id == sec.SectionID).count() if hasattr(Homework, 'section_id') else 0
    exams_count = ExamSchedule.query.filter(ExamSchedule.SectionID == sec.SectionID).count() if hasattr(ExamSchedule, 'SectionID') else 0

    linked_deps = []
    if students_count > 0: linked_deps.append(f"{students_count} طلاب")
    if timetable_count > 0: linked_deps.append(f"{timetable_count} حصص بالجدول")
    if homework_count > 0: linked_deps.append(f"{homework_count} واجبات")
    if exams_count > 0: linked_deps.append(f"{exams_count} امتحانات")

    if linked_deps:
        return jsonify({
            'success': False,
            'message': f'تعذر حذف الشعبة "{sec.SectionName}" لارتباطها بـ ({", ".join(linked_deps)}) لصف {c.CName}.'
        }), 400

    from models.academic import ClassesSections
    db.session.execute(ClassesSections.delete().where(ClassesSections.c.CID == class_id, ClassesSections.c.SectionID == sec_id))
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم حذف الشعبة "{sec.SectionName}" بنجاح من صف {c.CName}'})

@academic_bp.route('/add_subject', methods=['POST'])
def add_subject():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    from models.teacher import Teacher
    name = request.form.get('name', '').strip()
    sub_type = request.form.get('type', 'أساسية')
    department = request.form.get('department', 'جميع المراحل')
    weekly_hours = request.form.get('weekly_hours', type=int) or 4
    status = request.form.get('status', 'نشط')
    color = request.form.get('color', '#2563eb')
    class_ids = request.form.getlist('class_ids')
    teacher_ids = request.form.getlist('teacher_ids')
    
    if not name:
        flash('يرجى إدخال اسم المادة الدراسية', 'warning')
        return redirect(url_for('academic.subjects'))
        
    from sqlalchemy import func
    from sqlalchemy.exc import IntegrityError
    existing_subject = Subject.query.filter(
        func.lower(Subject.SubName) == name.lower(),
        Subject.is_deleted == False if hasattr(Subject, 'is_deleted') else True
    ).first()
    if existing_subject:
        flash(f'المادة الدراسية "{name}" موجودة بالفعل في النظام', 'warning')
        return redirect(url_for('academic.subjects'))
        
    try:
        subject_obj = Subject(
            SubName=name, 
            Type=sub_type, 
            Department=department, 
            WeeklyHours=weekly_hours, 
            Status=status, 
            Color=color
        )
        db.session.add(subject_obj)
        db.session.commit()
        flash(f'تمت إضافة المادة الدراسية "{name}" بنجاح', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'المادة الدراسية "{name}" موجودة بالفعل في النظام', 'warning')
        return redirect(url_for('academic.subjects'))

    if class_ids:
        subject_obj.classes = []
        target_classes = Classes.query.filter(Classes.CID.in_([int(cid) for cid in class_ids])).all()
        subject_obj.classes.extend(target_classes)

    if teacher_ids:
        subject_obj.teachers = []
        target_teachers = Teacher.query.filter(Teacher.TeacherID.in_([int(tid) for tid in teacher_ids])).all()
        subject_obj.teachers.extend(target_teachers)

    try:
        db.session.commit()
        from services.timetable_sync_service import sync_subject_timetable_slots
        sync_subject_timetable_slots(subject_obj.SubID)
        flash(f'تمت إضافة المادة "{name}" وتخصيص الصفوف والمعلمين بنجاح', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'المادة الدراسية "{name}" موجودة بالفعل أو حدث تعارض في البيانات', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة المادة: {str(e)}', 'danger')
        
    return redirect(url_for('academic.subjects'))

@academic_bp.route('/add_day', methods=['POST'])
def add_day():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    name = request.form.get('name')
    if name:
        try:
            db.session.add(Days(DName=name))
            db.session.commit()
            flash('تمت إضافة اليوم بنجاح', 'success')
        except Exception:
            db.session.rollback()
            flash('تعذر إضافة اليوم', 'danger')
    return redirect(url_for('academic.index'))

@academic_bp.route('/add_lesson', methods=['POST'])
def add_lesson():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    name = request.form.get('name')
    if name:
        try:
            db.session.add(Lessons(LessonName=name))
            db.session.commit()
            flash('تمت إضافة الحصة بنجاح', 'success')
        except Exception:
            db.session.rollback()
            flash('تعذر إضافة الحصة', 'danger')
    return redirect(url_for('academic.index'))

@academic_bp.route('/add_term', methods=['POST'])
def add_term():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    name = request.form.get('name')
    if name:
        try:
            db.session.add(Terms(T_Name=name))
            db.session.commit()
            flash('تمت إضافة الترم بنجاح', 'success')
        except Exception:
            db.session.rollback()
            flash('تعذر إضافة الترم', 'danger')
    return redirect(url_for('academic.index'))

@academic_bp.route('/add_exam_type', methods=['POST'])
def add_exam_type():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    name = request.form.get('name')
    if name:
        try:
            db.session.add(TypeExams(ExamName=name))
            db.session.commit()
            flash('تمت إضافة نوع الاختبار بنجاح', 'success')
        except Exception:
            db.session.rollback()
            flash('تعذر إضافة نوع الاختبار', 'danger')
    return redirect(url_for('academic.index'))


# --- EDIT AND DELETE ROUTES ---

def handle_edit(obj, attr_name, new_val):
    if new_val:
        setattr(obj, attr_name, new_val)
        try:
            db.session.commit()
            flash('تم التعديل بنجاح', 'success')
        except Exception:
            db.session.rollback()
            flash('تعذر التعديل بسبب تعارض في البيانات', 'danger')

def handle_delete(obj):
    try:
        db.session.delete(obj)
        db.session.commit()
        flash('تم الحذف بنجاح', 'success')
    except IntegrityError:
        db.session.rollback()
        flash('لا يمكن الحذف لارتباط هذا العنصر ببيانات أخرى', 'danger')

# Classes
@academic_bp.route('/edit_class/<int:id>', methods=['POST'])
def edit_class(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    c = Classes.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    stage = request.form.get('stage', '').strip()
    if name: c.CName = name
    if stage: c.Stage = stage
    try:
        db.session.commit()
        flash('تم تعديل الصف بنجاح', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'اسم الصف "{name}" مستخدم بالفعل لصف آخر', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء التعديل: {str(e)}', 'danger')
    return redirect(url_for('academic.classes'))

@academic_bp.route('/delete_class/<int:id>', methods=['POST'])
def delete_class(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    c = Classes.query.get_or_404(id)
    student_count = Student.query.filter_by(CID=id, is_deleted=False).count()
    if student_count > 0:
        flash('لا يمكن حذف الصف لاحتوائه على طلاب مسجلين بالفعل. يرجى نقل الطلاب أو تفريغ الصف أولاً.', 'danger')
        return redirect(url_for('academic.classes'))
        
    try:
        c.is_deleted = True
        associated_sections = list(c.sections)
        c.sections.clear()
        for sec in associated_sections:
            remaining_classes = [cl for cl in sec.classes if not getattr(cl, 'is_deleted', False)]
            if not remaining_classes:
                sec.is_deleted = True
                
        db.session.commit()
        flash(f'تم حذف الصف "{c.CName}" والشعب المرتبطة به بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الصف: {str(e)}', 'danger')
        
    return redirect(url_for('academic.classes'))

# Sections
@academic_bp.route('/edit_section/<int:id>', methods=['POST'])
def edit_section(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    s = Sections.query.get_or_404(id)
    name = request.form.get('name')
    if name: s.SectionName = name
    class_id = request.form.get('class_id')
    if class_id:
        c = Classes.query.get(class_id)
        if c and c not in s.classes:
            s.classes.append(c)
    db.session.commit()
    flash('تم تعديل الشعبة بنجاح', 'success')
    return redirect(url_for('academic.classes'))

@academic_bp.route('/delete_section/<int:id>', methods=['POST'])
def delete_section(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    s = Sections.query.get_or_404(id)
    student_count = Student.query.filter_by(SectionID=id, is_deleted=False).count()
    if student_count > 0:
        flash('لا يمكن حذف الشعبة لاحتوائها على طلاب مسجلين بالفعل. يرجى نقل الطلاب أو تفريغ الشعبة أولاً.', 'danger')
        return redirect(url_for('academic.classes'))
    try:
        s.is_deleted = True
        s.classes.clear()
        db.session.commit()
        flash(f'تم حذف الشعبة "{s.SectionName}" بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الشعبة: {str(e)}', 'danger')
        
    return redirect(url_for('academic.classes'))

# Subjects (handled via enriched endpoints at end of module)

# Days
@academic_bp.route('/edit_day/<int:id>', methods=['POST'])
def edit_day(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    d = Days.query.get_or_404(id)
    handle_edit(d, 'DName', request.form.get('name'))
    return redirect(url_for('academic.index'))

@academic_bp.route('/delete_day/<int:id>', methods=['POST'])
def delete_day(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    d = Days.query.get_or_404(id)
    handle_delete(d)
    return redirect(url_for('academic.index'))

# Lessons
@academic_bp.route('/edit_lesson/<int:id>', methods=['POST'])
def edit_lesson(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    l = Lessons.query.get_or_404(id)
    handle_edit(l, 'LessonName', request.form.get('name'))
    return redirect(url_for('academic.index'))

@academic_bp.route('/delete_lesson/<int:id>', methods=['POST'])
def delete_lesson(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    l = Lessons.query.get_or_404(id)
    handle_delete(l)
    return redirect(url_for('academic.index'))

# Terms
@academic_bp.route('/edit_term/<int:id>', methods=['POST'])
def edit_term(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    t = Terms.query.get_or_404(id)
    handle_edit(t, 'T_Name', request.form.get('name'))
    return redirect(url_for('academic.index'))

@academic_bp.route('/delete_term/<int:id>', methods=['POST'])
def delete_term(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    t = Terms.query.get_or_404(id)
    handle_delete(t)
    return redirect(url_for('academic.index'))

# Exam Types
@academic_bp.route('/edit_exam_type/<int:id>', methods=['POST'])
def edit_exam_type(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    et = TypeExams.query.get_or_404(id)
    handle_edit(et, 'ExamName', request.form.get('name'))
    return redirect(url_for('academic.index'))

@academic_bp.route('/delete_exam_type/<int:id>', methods=['POST'])
def delete_exam_type(id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    et = TypeExams.query.get_or_404(id)
    handle_delete(et)
    return redirect(url_for('academic.index'))

@academic_bp.route('/classes/export/excel')
def export_classes_excel():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from flask import send_file
    
    ids_param = request.args.get('ids', '').strip()
    query = Classes.query.filter_by(is_deleted=False)
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        classes_list = query.filter(Classes.CID.in_(id_list)).all()
    else:
        classes_list = query.order_by(Classes.CID.asc()).all()
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قائمة الصفوف والشعب"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    headers = ["كود الصف", "اسم الصف", "المرحلة الدراسية", "عدد الشعب", "عدد الطلاب", "السعة القصوى", "نسبة الإشغال", "عدد المعلمين", "عدد المواد"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for c in classes_list:
        st_count = Student.query.filter_by(CID=c.CID, is_deleted=False).count()
        t_count = _get_class_teachers_count(c.CID)
        sub_count = len([sub for sub in c.subjects if not getattr(sub, 'is_deleted', False)])
        sec_count = len([sec for sec in c.sections if not getattr(sec, 'is_deleted', False)])
        
        occ = f"{round((st_count / c.MaxStudents) * 100, 1)}%" if c.MaxStudents and c.MaxStudents > 0 else "غير محددة"
        max_st = c.MaxStudents if c.MaxStudents else "غير محددة"
        
        row = [
            f"CLS-{c.CID}",
            c.CName,
            c.Stage or 'غير محددة',
            sec_count,
            st_count,
            max_st,
            occ,
            t_count,
            sub_count
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
            
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 20
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='classes_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@academic_bp.route('/classes/export/pdf')
def export_classes_pdf():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from datetime import datetime
    
    ids_param = request.args.get('ids', '').strip()
    query = Classes.query.filter_by(is_deleted=False)
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        classes_list = query.filter(Classes.CID.in_(id_list)).all()
    else:
        classes_list = query.order_by(Classes.CID.asc()).all()
        
    for c in classes_list:
        c.linked_sections = [s for s in c.sections if not getattr(s, 'is_deleted', False)]
        c.students_count = Student.query.filter_by(CID=c.CID, is_deleted=False).count()
        c.teachers_count = _get_class_teachers_count(c.CID)
        c.subjects_count = len([sub for sub in c.subjects if not getattr(sub, 'is_deleted', False)])
        c.occ_str = f"{round((c.students_count / c.MaxStudents) * 100, 1)}%" if c.MaxStudents and c.MaxStudents > 0 else "غير محددة"
        
    return render_template('academic/classes_pdf_report.html', classes=classes_list, generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'))

@academic_bp.route('/classes/bulk-delete', methods=['POST'])
def bulk_delete_classes():
    from flask import jsonify
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'الرجاء تسجيل الدخول أولاً'}), 401
        
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد أي صف للحذف'}), 400
            
        blocked_classes = []
        classes_to_delete = []
        
        for cid in ids:
            c = Classes.query.filter_by(CID=cid, is_deleted=False).first()
            if c:
                st_count = Student.query.filter_by(CID=c.CID, is_deleted=False).count()
                if st_count > 0:
                    blocked_classes.append(c.CName)
                else:
                    classes_to_delete.append(c)
                    
        if blocked_classes:
            msg = f"تعذر حذف الصفوف التالية لاحتوائها على طلاب مسجلين: ({', '.join(blocked_classes)}). يرجى نقل الطلاب أولاً."
            return jsonify({'success': False, 'message': msg, 'blocked': True})
            
        if not classes_to_delete:
            return jsonify({'success': False, 'message': 'لم يتم العثور على صفوف قابلة للحذف'}), 400

        for c in classes_to_delete:
            c.is_deleted = True
            associated_sections = list(c.sections)
            c.sections.clear()
            for sec in associated_sections:
                remaining_classes = [cl for cl in sec.classes if not getattr(cl, 'is_deleted', False)]
                if not remaining_classes:
                    sec.is_deleted = True
            
        db.session.commit()
        return jsonify({'success': True, 'message': f'تم حذف {len(classes_to_delete)} صفوف بنجاح', 'count': len(classes_to_delete)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ غير متوقع أثناء الحذف: {str(e)}'}), 500

@academic_bp.route('/classes/bulk-status', methods=['POST'])
def bulk_status_classes():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'الرجاء تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    data = request.get_json() or {}
    ids = data.get('ids', [])
    new_status = data.get('status', 'نشط')
    if not ids:
        return jsonify({'success': False, 'message': 'لم يتم تحديد أي صف لتحديث الحالة'}), 400
        
    classes_list = Classes.query.filter(Classes.CID.in_(ids)).all()
    for c in classes_list:
        if hasattr(c, 'Status'):
            c.Status = new_status
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم تحديث حالة {len(classes_list)} صفوف إلى "{new_status}" بنجاح', 'count': len(classes_list)})

@academic_bp.route('/edit_subject/<int:id>', methods=['POST'])
def edit_subject(id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    from models.teacher import Teacher
    subject = Subject.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    sub_type = request.form.get('type')
    department = request.form.get('department')
    weekly_hours = request.form.get('weekly_hours', type=int) or 0
    status = request.form.get('status', 'نشط')
    color = request.form.get('color')
    class_ids = request.form.getlist('class_ids')
    teacher_ids = request.form.getlist('teacher_ids')
    
    if name:
        subject.SubName = name
        subject.Type = sub_type
        subject.Department = department
        subject.WeeklyHours = weekly_hours
        subject.Status = status
        if color:
            subject.Color = color
            
        if class_ids is not None:
            subject.classes = []
            if class_ids:
                target_classes = Classes.query.filter(Classes.CID.in_([int(cid) for cid in class_ids])).all()
                subject.classes.extend(target_classes)
                
        if teacher_ids is not None:
            subject.teachers = []
            if teacher_ids:
                target_teachers = Teacher.query.filter(Teacher.TeacherID.in_([int(tid) for tid in teacher_ids])).all()
                subject.teachers.extend(target_teachers)

        db.session.commit()
        from services.timetable_sync_service import sync_subject_timetable_slots
        sync_subject_timetable_slots(subject.SubID)
        flash('تم تحديث بيانات المادة بنجاح', 'success')
    return redirect(url_for('academic.subjects'))

@academic_bp.route('/delete_subject/<int:id>', methods=['POST'])
def delete_subject(id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    subject = Subject.query.get_or_404(id)
    
    from models import SchoolTable
    slots_count = SchoolTable.query.filter_by(SubID=subject.SubID, is_deleted=False).count()
    
    if slots_count > 0:
        flash(f'تعذر حذف المادة "{subject.SubName}" لارتباطها بـ {slots_count} حصص في الجدول الأسبوعي.', 'danger')
        return redirect(url_for('academic.subjects'))
        
    try:
        # Clear associations from join tables (TeacherSubject, ClassSubject)
        from models.academic import TeacherSubject, ClassSubject
        db.session.execute(TeacherSubject.delete().where(TeacherSubject.c.SubID == subject.SubID))
        db.session.execute(ClassSubject.delete().where(ClassSubject.c.SubID == subject.SubID))
        
        # Soft delete subject to preserve relational integrity across MySQL FK constraints (e.g. homework, exams)
        if hasattr(subject, 'is_deleted'):
            subject.is_deleted = True
        else:
            db.session.delete(subject)
            
        db.session.commit()
        flash(f'تم حذف المادة "{subject.SubName}" وإلغاء إسنادها بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ غير متوقع أثناء حذف المادة: {str(e)}', 'danger')

    return redirect(url_for('academic.subjects'))

@academic_bp.route('/subject/<int:id>/data')
def get_subject_data_api(id):
    from flask import jsonify
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from models.teacher import Teacher
    from models.grade import Marks
    subject = Subject.query.get_or_404(id)
    
    class_objs = subject.classes.all() if hasattr(subject.classes, 'all') else subject.classes
    linked_classes = []
    for c in class_objs:
        if not getattr(c, 'is_deleted', False):
            c_students = Student.query.filter_by(CID=c.CID, is_deleted=False).count()
            c_sections = len([sec for sec in c.sections if not getattr(sec, 'is_deleted', False)])
            c_occ = round((c_students / c.MaxStudents) * 100, 1) if c.MaxStudents and c.MaxStudents > 0 else None
            linked_classes.append({
                'id': c.CID,
                'name': c.CName,
                'stage': c.Stage or 'المرحلة العامة',
                'sectionsCount': c_sections,
                'studentsCount': c_students,
                'maxStudents': c.MaxStudents or 0,
                'occupancy': c_occ
            })

    teacher_objs = subject.teachers.all() if hasattr(subject.teachers, 'all') else subject.teachers
    assigned_teachers = [{'id': t.TeacherID, 'name': t.TeacherName, 'title': t.TeacherTitle or 'معلم قدير', 'email': t.Email or 'غير محدد', 'phone': t.Phone or 'غير محدد', 'status': t.Status or 'نشط', 'image': t.Image or ''} for t in teacher_objs if not getattr(t, 'is_deleted', False)]
    
    marks = Marks.query.filter_by(SubID=subject.SubID).all()
    if marks:
        valid_scores = [float(m.Score) for m in marks if m.Score is not None]
        avg_score = round(sum(valid_scores) / len(valid_scores), 1) if valid_scores else None
        passed_marks = [m for m in marks if m.Score is not None and (float(m.Score) / float(m.MaxScore or 100)) >= 0.5]
        pass_rate = round((len(passed_marks) / len(valid_scores)) * 100, 1) if valid_scores else None
    else:
        avg_score = None
        pass_rate = None

    class_ids = [c['id'] for c in linked_classes]
    students_count = Student.query.filter(Student.CID.in_(class_ids), Student.is_deleted == False).count() if class_ids else 0
    weekly_slots_count = SchoolTable.query.filter_by(SubID=subject.SubID, is_deleted=False).count()
    sections_count = sum(c['sectionsCount'] for c in linked_classes)

    # Real timetable slots from SchoolTable
    timetable_records = SchoolTable.query.filter_by(SubID=subject.SubID, is_deleted=False).all()
    slots_list = []
    for slot in timetable_records:
        slots_list.append({
            'day': slot.day.DName if slot.day else 'غير محدد',
            'lesson': slot.lesson.LessonName if slot.lesson else 'غير محدد',
            'className': slot.school_class.CName if slot.school_class else 'غير محدد',
            'sectionName': slot.section.SectionName if slot.section else ''
        })

    return jsonify({
        'success': True,
        'subject': {
            'id': subject.SubID,
            'code': f"SUB-{subject.SubID}",
            'name': subject.SubName,
            'type': subject.Type or 'أساسية',
            'department': subject.Department or 'جميع المراحل',
            'weeklyHours': subject.WeeklyHours or 4,
            'status': subject.Status or 'نشط',
            'color': subject.Color or '#2563eb',
            'studentsCount': students_count,
            'teachersCount': len(assigned_teachers),
            'classesCount': len(linked_classes),
            'sectionsCount': sections_count,
            'weeklySlotsCount': weekly_slots_count,
            'avgScore': avg_score,
            'passRate': pass_rate,
            'classes': linked_classes,
            'linkedClasses': linked_classes,
            'teachers': assigned_teachers,
            'classIds': class_ids,
            'teacherIds': [t['id'] for t in assigned_teachers],
            'timetableSlots': slots_list
        }
    })

@academic_bp.route('/subject/<int:id>/teachers', methods=['POST'])
def update_subject_teachers_api(id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    from models.teacher import Teacher
    subject = Subject.query.get_or_404(id)
    
    data = request.get_json() or {}
    teacher_ids = data.get('teacher_ids', [])
    
    subject.teachers = []
    if teacher_ids:
        target_teachers = Teacher.query.filter(Teacher.TeacherID.in_([int(tid) for tid in teacher_ids])).all()
        subject.teachers.extend(target_teachers)
        
    db.session.commit()
    teacher_objs = subject.teachers.all() if hasattr(subject.teachers, 'all') else subject.teachers
    return jsonify({'success': True, 'message': f'تم تحديث الكادر التعليمي للمادة "{subject.SubName}" بنجاح', 'assigned_count': len(teacher_objs)})

@academic_bp.route('/subject/<int:id>/classes', methods=['POST'])
def update_subject_classes_api(id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    subject = Subject.query.get_or_404(id)
    
    data = request.get_json() or {}
    class_ids = data.get('class_ids', [])
    
    subject.classes = []
    if class_ids:
        target_classes = Classes.query.filter(Classes.CID.in_([int(cid) for cid in class_ids])).all()
        subject.classes.extend(target_classes)
        
    db.session.commit()
    class_objs = subject.classes.all() if hasattr(subject.classes, 'all') else subject.classes
    return jsonify({'success': True, 'message': f'تم تحديث الصفوف المرتبطة بالمادة "{subject.SubName}" بنجاح', 'classes_count': len(class_objs)})

@academic_bp.route('/subjects/export/excel')
def export_subjects_excel():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from flask import send_file
    from models import SchoolTable
    from models.academic import TeacherSubject
    
    ids_param = request.args.get('ids', '').strip()
    query = Subject.query
    if hasattr(Subject, 'is_deleted'):
        query = query.filter_by(is_deleted=False)
        
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        subjects_list = query.filter(Subject.SubID.in_(id_list)).all()
    else:
        subjects_list = query.order_by(Subject.SubID.asc()).all()
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قائمة المواد الدراسية"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    headers = ["كود المادة", "اسم المادة", "النوع", "القسم/المرحلة", "الحصص الأسبوعية", "الصفوف المرتبطة", "عدد المعلمين", "عدد الطلاب", "الحالة"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for s in subjects_list:
        linked_cls = [c for c in s.classes if not getattr(c, 'is_deleted', False)]
        cls_names = ", ".join([c.CName for c in linked_cls]) or "جميع الصفوف"
        
        t_ids_ts = [t[0] for t in db.session.query(TeacherSubject.c.TeacherID).filter(TeacherSubject.c.SubID == s.SubID).distinct().all() if t[0]]
        t_ids_st = [t[0] for t in db.session.query(SchoolTable.TeacherID).filter(SchoolTable.SubID == s.SubID, SchoolTable.is_deleted == False).distinct().all() if t[0]]
        teachers_count = len(set(t_ids_ts + t_ids_st))
        
        class_ids = [c.CID for c in linked_cls]
        st_count = Student.query.filter(Student.CID.in_(class_ids), Student.is_deleted == False).count() if class_ids else 0
        
        row = [
            f"SUB-{s.SubID}",
            s.SubName,
            s.Type or 'أساسية',
            s.Department or 'جميع المراحل',
            getattr(s, 'WeeklyHours', 0) or 0,
            cls_names,
            teachers_count,
            st_count,
            getattr(s, 'Status', 'نشط') or 'نشط'
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
            
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 22
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='subjects_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@academic_bp.route('/subjects/export/pdf')
def export_subjects_pdf():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    from datetime import datetime
    from models import SchoolTable
    from models.academic import TeacherSubject
    
    ids_param = request.args.get('ids', '').strip()
    query = Subject.query
    if hasattr(Subject, 'is_deleted'):
        query = query.filter_by(is_deleted=False)
        
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        subjects_list = query.filter(Subject.SubID.in_(id_list)).all()
    else:
        subjects_list = query.order_by(Subject.SubID.asc()).all()
        
    for s in subjects_list:
        s.linked_classes = [c for c in s.classes if not getattr(c, 'is_deleted', False)]
        t_ids_ts = [t[0] for t in db.session.query(TeacherSubject.c.TeacherID).filter(TeacherSubject.c.SubID == s.SubID).distinct().all() if t[0]]
        t_ids_st = [t[0] for t in db.session.query(SchoolTable.TeacherID).filter(SchoolTable.SubID == s.SubID, SchoolTable.is_deleted == False).distinct().all() if t[0]]
        s.teachers_count = len(set(t_ids_ts + t_ids_st))
        class_ids = [c.CID for c in s.linked_classes]
        s.students_count = Student.query.filter(Student.CID.in_(class_ids), Student.is_deleted == False).count() if class_ids else 0
        s.weekly_slots_count = SchoolTable.query.filter(SchoolTable.SubID == s.SubID, SchoolTable.is_deleted == False).count()
        
    return render_template('academic/subjects_pdf_report.html', subjects=subjects_list, generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'))

@academic_bp.route('/subjects/bulk-status', methods=['POST'])
def bulk_status_subjects():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'الرجاء تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    data = request.get_json() or {}
    ids = data.get('ids', [])
    new_status = data.get('status', 'نشط')
    if not ids:
        return jsonify({'success': False, 'message': 'لم يتم تحديد أي مادة لتحديث الحالة'}), 400
        
    subjects_list = Subject.query.filter(Subject.SubID.in_(ids)).all()
    for s in subjects_list:
        if hasattr(s, 'Status'):
            s.Status = new_status
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم تحديث حالة {len(subjects_list)} مواد إلى "{new_status}" بنجاح', 'count': len(subjects_list)})

@academic_bp.route('/subjects/bulk-delete', methods=['POST'])
def bulk_delete_subjects():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'الرجاء تسجيل الدخول أولاً'}), 401
    from flask import jsonify
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'لم يتم تحديد أي مادة للحذف'}), 400
        
    from models import SchoolTable
    blocked_subjects = []
    subjects_to_delete = []
    
    for sub_id in ids:
        s = Subject.query.get(sub_id)
        if s:
            slots_count = SchoolTable.query.filter_by(SubID=s.SubID, is_deleted=False).count()
            if slots_count > 0:
                blocked_subjects.append(s.SubName)
            else:
                subjects_to_delete.append(s)
                
    if blocked_subjects:
        msg = f"تعذر حذف المواد التالية لارتباطها بجدول الحصص الأسبوعي: ({', '.join(blocked_subjects)})."
        return jsonify({'success': False, 'message': msg, 'blocked': True})
        
    for s in subjects_to_delete:
        if hasattr(s, 'is_deleted'):
            s.is_deleted = True
        else:
            db.session.delete(s)
            
    db.session.commit()
    return jsonify({'success': True, 'message': f'تم حذف {len(subjects_to_delete)} مواد بنجاح', 'count': len(subjects_to_delete)})
