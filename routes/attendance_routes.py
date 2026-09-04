from flask import Blueprint, render_template, request, session, redirect, url_for, jsonify
from flask_login import login_required, current_user
from datetime import date, datetime
from sqlalchemy.orm import joinedload
from models import db, Student, Classes, Sections, Teacher, SchoolTable, Terms, Subject
from models.student import Attendance
from utils.decorators import admin_required
from services.teacher_attendance_service import get_lesson_attendance, save_lesson_attendance

attendance_bp = Blueprint('attendance', __name__, url_prefix='/attendance')

from sqlalchemy import or_

def get_teacher_attendance_data(user_id, class_id=None, section_id=None, target_date=None, subject_id=None, status_filter=None, search_query=None):
    today = target_date if target_date else date.today()
    now_dt = datetime.now()
    now_time_str = now_dt.strftime('%H:%M')
    is_today = (today == now_dt.date())
    is_past = (today < now_dt.date())
    is_future = (today > now_dt.date())
    arabic_days = ['الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']
    today_day_name = arabic_days[today.weekday()]

    teacher = Teacher.query.options(joinedload(Teacher.subjects)).filter_by(user_id=user_id).first()
    
    if not teacher:
        teacher_name = 'مدير النظام'
        teacher_title = 'إدارة النظام'
        teacher_subjects = Subject.query.filter_by(is_deleted=False).all()
        sub_names = [s.SubName for s in teacher_subjects]
        slots = SchoolTable.query.options(
            joinedload(SchoolTable.subject),
            joinedload(SchoolTable.school_class),
            joinedload(SchoolTable.section),
            joinedload(SchoolTable.day),
            joinedload(SchoolTable.lesson)
        ).filter(SchoolTable.is_deleted == False).all()
    else:
        teacher_name = teacher.TeacherName
        teacher_title = teacher.TeacherTitle or 'معلم أكاديمي'
        slots = SchoolTable.query.options(
            joinedload(SchoolTable.subject),
            joinedload(SchoolTable.school_class),
            joinedload(SchoolTable.section),
            joinedload(SchoolTable.day),
            joinedload(SchoolTable.lesson)
        ).filter_by(TeacherID=teacher.TeacherID, is_deleted=False).all()

        slot_subs = [s.subject for s in slots if s.subject and not s.subject.is_deleted]
        t_subs = list(teacher.subjects) if teacher.subjects else []
        all_subs_map = {}
        for sub in (t_subs + slot_subs):
            if sub and sub.SubID not in all_subs_map and not getattr(sub, 'is_deleted', False):
                all_subs_map[sub.SubID] = sub
        teacher_subjects = list(all_subs_map.values()) if all_subs_map else Subject.query.filter_by(is_deleted=False).all()
        sub_names = [s.SubName for s in teacher_subjects]

    today_display = f"{today_day_name} {today.strftime('%Y-%m-%d')}"

    subjects_str = " | ".join(sub_names) if sub_names else 'المواد الدراسية'

    selected_subid = None
    selected_sub_name = None
    if subject_id is not None and str(subject_id).strip() != '':
        try:
            selected_subid = int(subject_id)
            sub_obj = Subject.query.get(selected_subid)
            if sub_obj:
                selected_sub_name = sub_obj.SubName
        except (ValueError, TypeError):
            selected_subid = None

    selected_cid = None
    if class_id is not None and str(class_id).strip() != '':
        try:
            selected_cid = int(class_id)
        except (ValueError, TypeError):
            selected_cid = None

    selected_secid = None
    if section_id is not None and str(section_id).strip() != '':
        try:
            selected_secid = int(section_id)
        except (ValueError, TypeError):
            selected_secid = None

    if selected_subid:
        scoped_slots = [s for s in slots if s.SubID == selected_subid]
        if not scoped_slots:
            scoped_slots = slots
        teacher_class_ids = list(set([s.CID for s in scoped_slots if s.CID]))
        teacher_section_ids = list(set([s.SectionID for s in scoped_slots if s.SectionID]))
    else:
        scoped_slots = slots
        teacher_class_ids = list(set([s.CID for s in slots if s.CID]))
        teacher_section_ids = list(set([s.SectionID for s in slots if s.SectionID]))

    class_sec_map = {}
    for s in slots:
        if s.school_class and s.section:
            c_name = s.school_class.CName
            sec_name = s.section.SectionName
            if c_name not in class_sec_map:
                class_sec_map[c_name] = []
            if sec_name not in class_sec_map[c_name]:
                class_sec_map[c_name].append(sec_name)

    formatted_classes_sections = []
    for c_name, secs in class_sec_map.items():
        secs_str = "، ".join(secs)
        formatted_classes_sections.append(f"{c_name} ({secs_str})")

    assigned_scope_str = " | ".join(formatted_classes_sections) if formatted_classes_sections else "الرابع (شعبة أ)"

    filter_slots = scoped_slots
    if selected_cid:
        filter_slots = [s for s in filter_slots if s.CID == selected_cid]
    if selected_secid:
        filter_slots = [s for s in filter_slots if s.SectionID == selected_secid]

    today_slots = [s for s in filter_slots if s.day and s.day.DName == today_day_name]
    today_slots_sorted = sorted(today_slots, key=lambda s: (s.lesson.StartTime if (s.lesson and s.lesson.StartTime) else '00:00'))

    current_slot = None
    status = 'لا توجد حصص مجدولة اليوم ☕'
    status_code = 'none'
    badge_class = 'bg-secondary'
    icon = 'fa-solid fa-mug-hot text-warning'
    remaining_minutes = None

    if today_slots_sorted:
        if is_today:
            # 1. Live slot in progress
            live_slot = None
            for s in today_slots_sorted:
                st_t = s.lesson.StartTime if (s.lesson and s.lesson.StartTime) else '00:00'
                en_t = s.lesson.EndTime if (s.lesson and s.lesson.EndTime) else '23:59'
                if st_t <= now_time_str <= en_t:
                    live_slot = s
                    break

            if live_slot:
                current_slot = live_slot
                status_code = 'current'
                status = 'الحصة جارية الآن 🟢'
                badge_class = 'bg-success'
                icon = 'fa-solid fa-circle-dot fa-beat text-white'
                try:
                    en_t = current_slot.lesson.EndTime
                    en_h, en_m = map(int, en_t.split(':'))
                    end_dt = datetime(now_dt.year, now_dt.month, now_dt.day, en_h, en_m)
                    diff_mins = max(1, int((end_dt - now_dt).total_seconds() / 60))
                    remaining_minutes = f"متبقي {diff_mins} دقيقة لنهاية الحصة ({en_t})"
                except Exception:
                    remaining_minutes = f"تنتهي الحصة في تمام {current_slot.lesson.EndTime}"
            else:
                # 2. Upcoming slots today
                upcoming_slots = [s for s in today_slots_sorted if (s.lesson and s.lesson.StartTime and s.lesson.StartTime > now_time_str)]
                if upcoming_slots:
                    current_slot = upcoming_slots[0]
                    status_code = 'upcoming'
                    status = 'الحصة القادمة ⏳'
                    badge_class = 'bg-primary'
                    icon = 'fa-solid fa-hourglass-start text-warning'
                    try:
                        st_t = current_slot.lesson.StartTime
                        st_h, st_m = map(int, st_t.split(':'))
                        start_dt = datetime(now_dt.year, now_dt.month, now_dt.day, st_h, st_m)
                        diff_mins = max(1, int((start_dt - now_dt).total_seconds() / 60))
                        remaining_minutes = f"تبدأ بعد {diff_mins} دقيقة (الساعة {st_t})"
                    except Exception:
                        remaining_minutes = f"موعد الحصة القادمة: {current_slot.lesson.StartTime}"
                else:
                    # 3. All lessons today have ended
                    current_slot = today_slots_sorted[-1]
                    status_code = 'ended'
                    status = 'انتهت الحصة ⏱️'
                    badge_class = 'bg-secondary'
                    icon = 'fa-solid fa-clock-rotate-left text-light'
                    remaining_minutes = f"انتهت الحصة في تمام الساعة {current_slot.lesson.EndTime}"

        elif is_past:
            current_slot = today_slots_sorted[0]
            status_code = 'ended'
            status = 'حصة سابقة (منتهية) 📅'
            badge_class = 'bg-secondary'
            icon = 'fa-solid fa-calendar-check text-light'
            remaining_minutes = f"حصة مسجلة بتاريخ {today.strftime('%Y-%m-%d')}"

        elif is_future:
            current_slot = today_slots_sorted[0]
            status_code = 'upcoming'
            status = 'حصة مجدولة قادمة ⏳'
            badge_class = 'bg-primary'
            icon = 'fa-solid fa-calendar-days text-info'
            remaining_minutes = f"مجدولة في موعدها ({current_slot.lesson.StartTime} - {current_slot.lesson.EndTime})"

    else:
        # No slots today (e.g. Friday / weekend)
        current_slot = None
        status_code = 'none'
        if today_day_name in ['الجمعة', 'السبت']:
            status = 'عطلة نهاية الأسبوع 🌴'
        else:
            status = 'لا توجد حصص مجدولة اليوم ☕'
        badge_class = 'bg-secondary'
        icon = 'fa-solid fa-mug-hot text-warning'
        remaining_minutes = 'لا توجد حصص في جدول اليوم'

    # Determine timing string
    if current_slot and current_slot.lesson:
        cur_st_time = current_slot.lesson.StartTime or 'غير محدد'
        cur_en_time = current_slot.lesson.EndTime or 'غير محدد'
        lesson_time_display = f"{cur_st_time} - {cur_en_time}"
    else:
        if filter_slots:
            first_slot = sorted(filter_slots, key=lambda s: (s.day.DayID if s.day else 0, s.lesson.StartTime if s.lesson else '00:00'))[0]
            day_text = first_slot.day.DName if first_slot.day else 'الأحد'
            st_text = first_slot.lesson.StartTime if first_slot.lesson else '08:00'
            en_text = first_slot.lesson.EndTime if first_slot.lesson else '08:45'
            lesson_time_display = f"أقرب حصة: {day_text} ({st_text} - {en_text})"
        else:
            lesson_time_display = "الدوام المدرسي المعتمد"

    # Determine Subject
    if selected_sub_name:
        cur_sub = selected_sub_name
    elif current_slot and current_slot.subject:
        cur_sub = current_slot.subject.SubName
    elif teacher_subjects:
        cur_sub = teacher_subjects[0].SubName
    else:
        cur_sub = 'العلوم'

    # Determine Class and Section
    cur_cls = None
    cur_sec = None
    if selected_cid:
        cls_obj = Classes.query.filter_by(CID=selected_cid).first()
        if cls_obj: cur_cls = cls_obj.CName
    elif current_slot and current_slot.school_class:
        cur_cls = current_slot.school_class.CName
    elif formatted_classes_sections:
        cur_cls = list(class_sec_map.keys())[0]
    else:
        cur_cls = 'الرابع'

    if selected_secid:
        sec_obj = Sections.query.filter_by(SectionID=selected_secid).first()
        if sec_obj: cur_sec = sec_obj.SectionName
    elif current_slot and current_slot.section:
        cur_sec = current_slot.section.SectionName
    elif formatted_classes_sections:
        cur_sec = class_sec_map.get(cur_cls, ['شعبة أ'])[0]
    else:
        cur_sec = 'شعبة أ'

    if selected_cid and selected_secid:
        class_sec_label = f"الصف: {cur_cls} - {cur_sec}"
    elif selected_cid and not selected_secid:
        class_sec_label = f"الصف: {cur_cls}"
    elif not selected_cid and selected_secid:
        class_sec_label = f"الشعبة: {cur_sec}"
    elif cur_cls and cur_sec:
        class_sec_label = f"الصف: {cur_cls} ({cur_sec})"
    elif formatted_classes_sections:
        class_sec_label = f"الصف: {formatted_classes_sections[0]}"
    else:
        class_sec_label = "الصف: الرابع (شعبة أ)"

    query = Student.query.options(joinedload(Student.school_class), joinedload(Student.section)).filter(Student.is_deleted == False, Student.CID.isnot(None))
    if teacher_class_ids:
        query = query.filter(Student.CID.in_(teacher_class_ids))
    if teacher_section_ids:
        query = query.filter(Student.SectionID.in_(teacher_section_ids))

    if selected_cid:
        query = query.filter_by(CID=selected_cid)
    if selected_secid:
        query = query.filter_by(SectionID=selected_secid)

    if search_query and str(search_query).strip():
        sq_str = str(search_query).strip()
        sq_pattern = f"%{sq_str}%"
        query = query.filter(or_(
            Student.SName.ilike(sq_pattern),
            db.cast(Student.SID, db.String).ilike(sq_pattern)
        ))

    students = query.all()
    student_ids = [s.SID for s in students]

    att_records = Attendance.query.filter(
        Attendance.SID.in_(student_ids), Attendance.Date == today
    ).order_by(
        Attendance.updated_at.desc(), Attendance.created_at.desc(), Attendance.AttendanceID.desc()
    ).all() if student_ids else []

    att_rec_map = {}
    for a in att_records:
        if a.SID not in att_rec_map:
            att_rec_map[a.SID] = a
    att_dict = {sid: a.Status for sid, a in att_rec_map.items()}

    present_c = sum(1 for status in att_dict.values() if status in ['Present', 'حاضر'])
    absent_c = sum(1 for status in att_dict.values() if status in ['Absent', 'غائب'])
    late_c = sum(1 for status in att_dict.values() if status in ['Late', 'متأخر', 'تأخر'])
    excused_c = sum(1 for status in att_dict.values() if status in ['Excused', 'مستأذن', 'بعذر'])
    
    total_st = len(students)
    unregistered_c = total_st - len(att_dict)

    if total_st > 0 and len(att_records) > 0:
        present_rate = round((present_c / total_st) * 100, 1)
        absent_rate = round((absent_c / total_st) * 100, 1)
        late_rate = round((late_c / total_st) * 100, 1)
        excused_rate = round((excused_c / total_st) * 100, 1)
        attendance_rate_display = f"{present_rate}%"
    else:
        present_rate = 0.0
        absent_rate = 0.0
        late_rate = 0.0
        excused_rate = 0.0
        attendance_rate_display = "لم يتم تسجيل الحضور بعد"

    att_rate = present_rate
    abs_rate = absent_rate
    disc_score = round(((present_c * 1.0 + late_c * 0.8 + excused_c * 0.9) / (total_st or 1)) * 100, 1)

    attendance_cards = []
    for idx, st in enumerate(students, start=1):
        rec = att_rec_map.get(st.SID)
        st_status = rec.Status if rec else 'غير مسجل'
        if st_status in ['Present', 'حاضر']:
            st_status_clean = 'حاضر'
            st_status_color = 'success'
        elif st_status in ['Absent', 'غائب']:
            st_status_clean = 'غائب'
            st_status_color = 'danger'
        elif st_status in ['Late', 'متأخر', 'تأخر']:
            st_status_clean = 'متأخر'
            st_status_color = 'warning'
        elif st_status in ['Excused', 'مستأذن', 'بعذر']:
            st_status_clean = 'بعذر'
            st_status_color = 'info'
        else:
            st_status_clean = 'غير مسجل'
            st_status_color = 'secondary'

        cls_n = st.school_class.CName if st.school_class else cur_cls
        sec_n = st.section.SectionName if st.section else cur_sec

        rec_time = '—'
        rec_dt = getattr(rec, 'updated_at', None) or getattr(rec, 'created_at', None) if rec else None
        if rec and rec_dt:
            h = rec_dt.strftime('%I').lstrip('0')
            if not h: h = '12'
            m = rec_dt.strftime('%M')
            s = rec_dt.strftime('%S')
            am_pm = 'ص' if rec_dt.strftime('%p') == 'AM' else 'م'
            rec_time = f"{h}:{m}:{s} {am_pm}"

        attendance_cards.append({
            'SID': st.SID,
            'SName': st.SName,
            'student_code': f"#{st.SID}",
            'class_name': cls_n,
            'section_name': sec_n,
            'subject_name': cur_sub,
            'status': st_status_clean,
            'status_color': st_status_color,
            'time_recorded': rec_time
        })

    # Apply status filter if specified
    if status_filter and str(status_filter).strip():
        sf_clean = str(status_filter).strip()
        attendance_cards = [c for c in attendance_cards if c['status'] == sf_clean or (sf_clean == 'بعذر' and c['status'] in ['بعذر', 'مستأذن'])]

    # Dynamic Most Absent
    most_absent = []
    for st in students:
        st_abs_cnt = db.session.query(db.func.count(Attendance.AttendanceID)).filter(Attendance.SID == st.SID, Attendance.Status.in_(['Absent', 'غائب'])).scalar() or 0
        if st_abs_cnt > 0 or att_dict.get(st.SID) in ['Absent', 'غائب']:
            effective_cnt = max(st_abs_cnt, 1 if att_dict.get(st.SID) in ['Absent', 'غائب'] else 0)
            most_absent.append({
                'SName': st.SName,
                'days': f"{effective_cnt} أيام",
                'count': effective_cnt
            })
    most_absent = sorted(most_absent, key=lambda x: x['count'], reverse=True)[:3]

    # Dynamic Attendance Alerts
    alerts = []
    if student_ids:
        absent_counts = db.session.query(
            Attendance.SID, db.func.count(Attendance.AttendanceID)
        ).filter(
            Attendance.SID.in_(student_ids),
            Attendance.Status.in_(['Absent', 'غائب'])
        ).group_by(Attendance.SID).all()
        
        absent_sid_map = {sid: count for sid, count in absent_counts}
        for st in students:
            if absent_sid_map.get(st.SID, 0) >= 3 or att_dict.get(st.SID) in ['Absent', 'غائب']:
                cls_n = st.school_class.CName if st.school_class else cur_cls
                sec_n = st.section.SectionName if st.section else cur_sec
                alerts.append({
                    'type': 'danger',
                    'title': 'طالب تجاوز حد الغياب المسموح',
                    'subtitle': f"{st.SName} - {sec_n} ({cls_n})"
                })
                break

        late_counts = db.session.query(
            Attendance.SID, db.func.count(Attendance.AttendanceID)
        ).filter(
            Attendance.SID.in_(student_ids),
            Attendance.Status.in_(['Late', 'متأخر', 'تأخر'])
        ).group_by(Attendance.SID).all()
        
        late_sid_map = {sid: count for sid, count in late_counts}
        for st in students:
            if late_sid_map.get(st.SID, 0) >= 2 or att_dict.get(st.SID) in ['Late', 'متأخر', 'تأخر']:
                cls_n = st.school_class.CName if st.school_class else cur_cls
                sec_n = st.section.SectionName if st.section else cur_sec
                alerts.append({
                    'type': 'warning',
                    'title': 'تكرر التأخير هذا الأسبوع',
                    'subtitle': f"{st.SName} - {sec_n} ({cls_n})"
                })
                break

    if total_st > 0:
        if len(att_records) == 0:
            alerts.append({
                'type': 'info',
                'title': 'شعبة لم يتم تسجيل حضورها اليوم',
                'subtitle': f"{cur_sec} - {cur_cls}"
            })
        else:
            alerts.append({
                'type': 'success',
                'title': 'تم تسجيل حضور هذه الشعبة اليوم',
                'subtitle': f"عدد المسجلين: {len(att_records)} من إجمالي {total_st} طالب"
            })

    kpi = {
        'total_students': total_st,
        'present_count': present_c,
        'absent_count': absent_c,
        'late_count': late_c,
        'excused_count': excused_c,
        'present_rate': present_rate,
        'absent_rate': absent_rate,
        'late_rate': late_rate,
        'excused_rate': excused_rate,
        'attendance_rate': att_rate,
        'absence_rate': abs_rate,
        'discipline_score': disc_score
    }

    current_lesson_info = {
        'subject': cur_sub,
        'time': lesson_time_display,
        'class_name': cur_cls,
        'section_name': cur_sec,
        'class_sec_label': class_sec_label,
        'lesson_num': getattr(current_slot.lesson, 'LName', None) if (current_slot and current_slot.lesson) else ('الحصة المجدولة' if current_slot else 'المقرر الأكاديمي'),
        'students_count': total_st,
        'remaining_minutes': remaining_minutes,
        'status': status,
        'status_code': status_code,
        'badge_class': badge_class,
        'icon': icon
    }

    teacher_info = {
        'TeacherName': teacher_name,
        'TeacherTitle': teacher_title,
        'subjects_str': subjects_str,
        'assigned_scope_str': assigned_scope_str
    }

    active_slot_id = current_slot.SchoolTableID if current_slot else (today_slots_sorted[0].SchoolTableID if today_slots_sorted else (slots[0].SchoolTableID if slots else None))

    return {
        'teacher_info': teacher_info,
        'current_lesson': current_lesson_info,
        'kpi': kpi,
        'attendance_cards': attendance_cards,
        'most_absent': most_absent,
        'alerts': alerts,
        'selected_cid': selected_cid,
        'selected_secid': selected_secid,
        'selected_subid': selected_subid,
        'teacher_subjects': teacher_subjects,
        'today_day_name': today_day_name,
        'today_display': today_display,
        'active_slot_id': active_slot_id
    }

@attendance_bp.route('/api/lesson/<int:slot_id>')
@login_required
def get_lesson_attendance_api(slot_id):
    date_str = request.args.get('date')
    data = get_lesson_attendance(slot_id, current_user.id, date_str=date_str)
    if not data:
        return jsonify({'error': 'Lesson attendance not found or access forbidden'}), 403
    return jsonify(data)

@attendance_bp.route('/api/save', methods=['POST'])
@login_required
def save_lesson_attendance_api():
    payload = request.json or {}
    slot_id = payload.get('slot_id')
    attendance_list = payload.get('attendance', [])
    date_str = payload.get('date_str')

    if not slot_id or not isinstance(attendance_list, list):
        return jsonify({'error': 'Invalid request payload'}), 400

    res = save_lesson_attendance(slot_id, current_user.id, attendance_list, date_str=date_str)
    if not res:
        return jsonify({'error': 'Failed to save attendance or access forbidden'}), 403

    return jsonify(res)

@attendance_bp.route('/')
@login_required
def index():
    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        date_str = request.args.get('date')
        class_id = request.args.get('class_id')
        section_id = request.args.get('section_id')
        subject_id = request.args.get('subject_id')
        status_filter = request.args.get('status')
        search_query = request.args.get('search')

        target_date = date.today()
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                target_date = date.today()
        
        teacher_class_ids = set()
        teacher_section_ids = set()
        if teacher:
            slots = SchoolTable.query.filter_by(TeacherID=teacher.TeacherID, is_deleted=False).all()
            for s in slots:
                if s.CID: teacher_class_ids.add(s.CID)
                if s.SectionID: teacher_section_ids.add(s.SectionID)

        if not teacher_class_ids:
            assigned_students = Student.query.filter(Student.is_deleted == False, Student.CID.isnot(None)).all()
            for st in assigned_students:
                if st.CID: teacher_class_ids.add(st.CID)
                if st.SectionID: teacher_section_ids.add(st.SectionID)

        classes = Classes.query.filter(Classes.CID.in_(teacher_class_ids), Classes.is_deleted == False).all() if teacher_class_ids else []

        if class_id and str(class_id).strip():
            try:
                cid_val = int(class_id)
                if teacher_class_ids and cid_val not in teacher_class_ids:
                    return jsonify({'error': 'Access to out-of-scope class forbidden'}), 403
                cls_obj = Classes.query.get(cid_val)
                if cls_obj and cls_obj.sections:
                    sec_list = [sec for sec in cls_obj.sections if not sec.is_deleted]
                    if teacher_section_ids:
                        sec_list = [sec for sec in sec_list if sec.SectionID in teacher_section_ids]
                    sections = sec_list if sec_list else (Sections.query.filter(Sections.SectionID.in_(teacher_section_ids), Sections.is_deleted == False).all() if teacher_section_ids else [])
                else:
                    sections = Sections.query.filter(Sections.SectionID.in_(teacher_section_ids), Sections.is_deleted == False).all() if teacher_section_ids else []
            except (ValueError, TypeError):
                sections = Sections.query.filter(Sections.SectionID.in_(teacher_section_ids), Sections.is_deleted == False).all() if teacher_section_ids else []
        else:
            sections = Sections.query.filter(Sections.SectionID.in_(teacher_section_ids), Sections.is_deleted == False).all() if teacher_section_ids else []
        
        data = get_teacher_attendance_data(
            current_user.id,
            class_id=class_id,
            section_id=section_id,
            target_date=target_date,
            subject_id=subject_id,
            status_filter=status_filter,
            search_query=search_query
        )

        active_slot_id = data.get('active_slot_id')

        return render_template('teacher/attendance.html',
                               classes=classes,
                               sections=sections,
                               teacher_subjects=data.get('teacher_subjects', []),
                               selected_cid=data.get('selected_cid'),
                               selected_secid=data.get('selected_secid'),
                               selected_subject_id=subject_id or '',
                               selected_status=status_filter or '',
                               search_query=search_query or '',
                               today=target_date.strftime('%Y-%m-%d'),
                               today_day_name=data.get('today_day_name', ''),
                               today_display=data.get('today_display', ''),
                               teacher_info=data['teacher_info'],
                               current_lesson=data['current_lesson'],
                               kpi=data['kpi'],
                               attendance_cards=data['attendance_cards'],
                               most_absent=data['most_absent'],
                               alerts=data['alerts'],
                               active_slot_id=active_slot_id)

    classes = Classes.query.filter_by(is_deleted=False).all()
    sections = Sections.query.filter_by(is_deleted=False).all()
    subjects = Subject.query.filter_by(is_deleted=False).all()
    terms = Terms.query.filter_by(is_deleted=False).all()
    
    class_id = request.args.get('class_id')
    section_id = request.args.get('section_id')
    date_str = request.args.get('date')
    subject_id = request.args.get('subject_id')
    term_id = request.args.get('term_id')

    target_date = date.today()
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            target_date = date.today()
    
    user_id = session.get('user_id', current_user.id if current_user.is_authenticated else 1)
    data = get_teacher_attendance_data(user_id, class_id=class_id, section_id=section_id, target_date=target_date)
    
    if subject_id:
        try:
            sub_obj = Subject.query.get(int(subject_id))
            if sub_obj:
                data['current_lesson']['subject'] = sub_obj.SubName
        except (ValueError, TypeError):
            pass

    return render_template('attendance.html',
                           classes=classes,
                           sections=sections,
                           subjects=subjects,
                           terms=terms,
                           today=target_date.strftime('%Y-%m-%d'),
                           teacher_info=data['teacher_info'],
                           current_lesson=data['current_lesson'],
                           kpi=data['kpi'],
                           attendance_cards=data['attendance_cards'],
                           most_absent=data['most_absent'],
                           alerts=data['alerts'],
                           selected_cid=data['selected_cid'],
                           selected_secid=data['selected_secid'],
                           selected_subject_id=subject_id or '',
                           selected_term_id=term_id or '',
                           total_students=data['kpi']['total_students'],
                           present=data['kpi']['present_count'],
                           absent=data['kpi']['absent_count'],
                           late=data['kpi']['late_count'],
                           attendance_rate=data['kpi']['attendance_rate'])

@attendance_bp.route('/api/students')
def get_students():
    class_id = request.args.get('class_id')
    section_id = request.args.get('section_id')
    target_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    
    query = Student.query.filter_by(Status='نشط', is_deleted=False)
    if class_id:
        query = query.filter_by(CID=class_id)
    if section_id:
        query = query.filter_by(SectionID=section_id)
        
    students = query.all()
    student_ids = [s.SID for s in students]
    
    attendances = {}
    if student_ids:
        att_records = Attendance.query.filter(Attendance.SID.in_(student_ids), Attendance.Date == target_date).all()
        attendances = {att.SID: att for att in att_records}
    
    result = []
    for s in students:
        att = attendances.get(s.SID)
        status = att.Status if att else 'غير مسجل'
        result.append({
            'SID': s.SID,
            'SName': s.SName,
            'Status': status,
            'Time': getattr(att, 'created_at', None),
            'Note': getattr(att, 'Note', '-')
        })
        
    return jsonify({'success': True, 'data': result})

@attendance_bp.route('/api/mark', methods=['POST'])
def mark_attendance():
    if 'user_id' not in session and not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}), 401
        
    data = request.json or {}
    sid = data.get('sid')
    target_date_str = data.get('date', date.today().strftime('%Y-%m-%d'))
    status = data.get('status')
    
    if not all([sid, target_date_str, status]):
        return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400
        
    try:
        sid = int(sid)
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'بيانات غير صالحة'}), 400

    if target_date > date.today():
        return jsonify({'success': False, 'message': 'لا يمكن تسجيل الحضور بتاريخ مستقبلي.'}), 400

    student = Student.query.filter_by(SID=sid, is_deleted=False).first()
    if not student:
        return jsonify({'success': False, 'message': 'الطالب غير موجود في النظام'}), 404
        
    now_dt = datetime.now()
    existing_records = Attendance.query.filter_by(SID=sid, Date=target_date).order_by(
        Attendance.updated_at.desc(), Attendance.created_at.desc(), Attendance.AttendanceID.desc()
    ).all()

    if existing_records:
        att = existing_records[0]
        att.Status = status
        att.updated_at = now_dt
        for dup in existing_records[1:]:
            db.session.delete(dup)
    else:
        att = Attendance(SID=sid, Date=target_date, Status=status, created_at=now_dt, updated_at=now_dt)
        db.session.add(att)
        
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': f'تم تسجيل الطالب كـ {status}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@attendance_bp.route('/export')
def export_attendance():
    if 'user_id' not in session and not getattr(current_user, 'is_authenticated', False):
        return redirect(url_for('auth.login'))
        
    class_id = request.args.get('class_id')
    section_id = request.args.get('section_id')
    subject_id = request.args.get('subject_id')
    target_date = request.args.get('date') or date.today().strftime('%Y-%m-%d')
    export_type = request.args.get('type')
    status_filter = request.args.get('status')
    only_recorded = request.args.get('only_recorded', type=int)
    sids_param = request.args.get('sids')

    user_id = session.get('user_id', current_user.id if getattr(current_user, 'is_authenticated', False) else 1)
    teacher = Teacher.query.options(joinedload(Teacher.subjects)).filter_by(user_id=user_id).first()
    
    teacher_class_ids = set()
    teacher_section_ids = set()
    teacher_sub_name = None

    if teacher:
        slots = SchoolTable.query.filter_by(TeacherID=teacher.TeacherID, is_deleted=False).all()
        for s in slots:
            if s.CID: teacher_class_ids.add(s.CID)
            if s.SectionID: teacher_section_ids.add(s.SectionID)
        if teacher.subjects:
            teacher_sub_name = teacher.subjects[0].SubName

    if not teacher_class_ids and teacher:
        assigned_students = Student.query.filter(Student.is_deleted == False, Student.CID.isnot(None)).all()
        for st in assigned_students:
            if st.CID: teacher_class_ids.add(st.CID)
            if st.SectionID: teacher_section_ids.add(st.SectionID)

    sub_name = 'جميع المواد'
    if subject_id and subject_id.isdigit():
        try:
            sub_obj = Subject.query.get(int(subject_id))
            if sub_obj: sub_name = sub_obj.SubName
        except (ValueError, TypeError):
            pass
    elif teacher_sub_name:
        sub_name = teacher_sub_name

    query = Student.query.options(joinedload(Student.school_class), joinedload(Student.section)).filter(Student.is_deleted == False, Student.CID.isnot(None))
    
    if class_id and class_id.isdigit():
        try:
            query = query.filter(Student.CID == int(class_id))
        except (ValueError, TypeError):
            pass
    elif teacher_class_ids:
        query = query.filter(Student.CID.in_(list(teacher_class_ids)))

    if section_id and section_id.isdigit():
        try:
            query = query.filter(Student.SectionID == int(section_id))
        except (ValueError, TypeError):
            pass
    elif teacher_section_ids and not class_id:
        query = query.filter(Student.SectionID.in_(list(teacher_section_ids)))

    if sids_param:
        try:
            sid_list = [int(s) for s in sids_param.split(',') if s.strip().isdigit()]
            if sid_list:
                query = query.filter(Student.SID.in_(sid_list))
        except Exception:
            pass

    students = query.order_by(Student.SID.asc()).all()
    student_ids = [s.SID for s in students]

    att_records = {}
    if student_ids:
        att_rows = Attendance.query.filter(Attendance.SID.in_(student_ids), Attendance.Date == target_date).all()
        att_records = {a.SID: a.Status for a in att_rows}

    # Filter export list based on requested status or recorded filter
    export_students = []
    for st in students:
        st_status = att_records.get(st.SID, 'لم يسجل')

        if only_recorded or status_filter in ['recorded', 'محضرين']:
            if st_status == 'لم يسجل':
                continue
        elif status_filter and status_filter in ['حاضر', 'غائب', 'متأخر', 'مستأذن']:
            if st_status != status_filter:
                continue
        elif status_filter in ['unrecorded', 'غير محضرين']:
            if st_status != 'لم يسجل':
                continue

        export_students.append((st, st_status))

    if export_type == 'pdf':
        cards = []
        for st, st_status in export_students:
            cards.append({
                'SID': st.SID,
                'SName': st.SName,
                'class_name': st.school_class.CName if st.school_class else '—',
                'section_name': st.section.SectionName if st.section else '—',
                'subject_name': sub_name,
                'status': st_status
            })
        return render_template('teacher/attendance_pdf.html',
                               students=cards,
                               today=target_date,
                               subject_name=sub_name,
                               generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'))

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    import io
    from flask import send_file
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف الحضور والغياب"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    headers = ["الرقم الطلابي", "اسم الطالب", "الصف والشعبة", "المادة الدراسية", "الحالة", "تاريخ الحضور"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for st, st_status in export_students:
        cls_sec = f"{st.school_class.CName if st.school_class else '—'} - {st.section.SectionName if st.section else '—'}"
        row = [st.SID, st.SName, cls_sec, sub_name, st_status, str(target_date)]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
            
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 25
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"attendance_export_{target_date}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

