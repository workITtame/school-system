from flask import Blueprint, render_template, request, make_response, send_file, current_app, jsonify, session
from flask_login import login_required, current_user
from models import db, Student, Classes, Sections, Marks, TypeExams, Subject, Terms, Teacher, Attendance, Homework
from models.extensions import cache
from utils.pdf_generator import generate_student_pdf
import io
import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

reports_bp = Blueprint('reports', __name__)

@reports_bp.route("/reports")
@reports_bp.route("/reports/")
@login_required
def index():
    from services.reports import get_reports_dashboard_metrics, get_reports_registry
    
    metrics = get_reports_dashboard_metrics()
    reports = get_reports_registry()
    
    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        from services.teacher_dashboard_service import get_teacher_by_user_id, get_teacher_subject_and_class_ids
        teacher = get_teacher_by_user_id(current_user.id)
        sub_ids, class_ids, sec_ids = get_teacher_subject_and_class_ids(teacher)
        classes = Classes.query.filter(Classes.CID.in_(class_ids), Classes.is_deleted == False).all() if class_ids else []
        sections = Sections.query.filter(Sections.SectionID.in_(sec_ids), Sections.is_deleted == False).all() if sec_ids else []
        subjects = Subject.query.filter(Subject.SubID.in_(sub_ids), Subject.is_deleted == False).all() if sub_ids else []
        students = Student.query.filter(Student.CID.in_(class_ids), Student.is_deleted == False).order_by(Student.SName).all() if class_ids else []
    else:
        classes = Classes.query.filter_by(is_deleted=False).all()
        sections = Sections.query.filter_by(is_deleted=False).all()
        subjects = Subject.query.filter_by(is_deleted=False).all()
        students = Student.query.filter_by(is_deleted=False).order_by(Student.SName).all()

    terms = Terms.query.filter_by(is_deleted=False).all()
    recent_reports = []

    return render_template("reports/index.html", 
                           metrics=metrics, 
                           reports=reports,
                           classes=classes,
                           sections=sections,
                           subjects=subjects,
                           terms=terms,
                           students=students,
                           recent_reports=recent_reports)

@reports_bp.route("/reports/analytics")
@login_required
def analytics():
    from services.reports import get_reports_dashboard_metrics
    metrics = get_reports_dashboard_metrics()
    return render_template("reports/analytics.html", metrics=metrics)

@reports_bp.route("/reports/student")
@login_required
def student_report():
    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        from services.teacher_dashboard_service import get_teacher_by_user_id, get_teacher_subject_and_class_ids
        teacher = get_teacher_by_user_id(current_user.id)
        sub_ids, class_ids, sec_ids = get_teacher_subject_and_class_ids(teacher)
        classes = Classes.query.filter(Classes.CID.in_(class_ids), Classes.is_deleted == False).all() if class_ids else []
        sections = Sections.query.filter(Sections.SectionID.in_(sec_ids), Sections.is_deleted == False).all() if sec_ids else []
    else:
        classes = Classes.query.filter_by(is_deleted=False).all()
        sections = Sections.query.filter_by(is_deleted=False).all()
    
    sel_class_id = request.args.get('class_id', type=int)
    sel_section_id = request.args.get('section_id', type=int)
    sel_student_id = request.args.get('student_id')
    
    query = Student.query.filter_by(is_deleted=False)
    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        if class_ids:
            query = query.filter(Student.CID.in_(class_ids))
        else:
            query = query.filter(Student.CID == -1)

    if sel_class_id:
        query = query.filter_by(CID=sel_class_id)
    if sel_section_id:
        query = query.filter_by(SectionID=sel_section_id)
    students = query.order_by(Student.SName).all()
        
    selected_student = None
    report_data = None
    average = 0
    
    if sel_student_id:
        try:
            s_id = int(sel_student_id)
            selected_student = Student.query.get(s_id)
        except (ValueError, TypeError):
            selected_student = None

        if selected_student:
            if hasattr(current_user, 'role') and current_user.role == 'teacher':
                from services.teacher_dashboard_service import get_teacher_by_user_id, get_teacher_subject_and_class_ids
                teacher = get_teacher_by_user_id(current_user.id)
                _, class_ids, _ = get_teacher_subject_and_class_ids(teacher)
                if not selected_student.CID or selected_student.CID not in class_ids:
                    return jsonify({'error': 'Out-of-scope student report access forbidden'}), 403
            marks = Marks.query.filter_by(SID=selected_student.SID, assessment_type='exam').all()
            report_data = {}
            total_score = 0
            count = 0
            for mark in marks:
                exam_name = "امتحان الفصل الدراسي"
                if hasattr(mark, 'exam') and mark.exam:
                    exam_name = mark.exam.ExamName
                elif mark.ExamID:
                    ex_obj = TypeExams.query.get(mark.ExamID)
                    if ex_obj: exam_name = ex_obj.ExamName

                sub_name = mark.subject.SubName if mark.subject else "مادة دراسية"
                score_val = float(mark.Score) if mark.Score is not None else 0.0

                grade_item = {
                    'subject_name': sub_name,
                    'score': score_val,
                    'grade_symbol': mark.Grade or '—',
                    'notes': mark.Notes or ''
                }

                if exam_name not in report_data:
                    report_data[exam_name] = []
                report_data[exam_name].append(grade_item)
                if mark.Score is not None:
                    total_score += score_val
                    count += 1
            if count > 0:
                average = round(total_score / count, 2)
                
    return render_template("reports/student_report.html",
                           classes=classes,
                           sections=sections,
                           students=students,
                           sel_class_id=sel_class_id,
                           sel_section_id=sel_section_id,
                           sel_student_id=sel_student_id,
                           selected_student=selected_student,
                           report_data=report_data,
                           average=average)

@reports_bp.route("/reports/performance")
@login_required
@cache.cached(timeout=60)
def performance():
    from sqlalchemy import func
    results = db.session.query(
        Classes.CName, 
        func.avg(Marks.Score)
    ).join(Student, Student.CID == Classes.CID)\
     .join(Marks, Marks.SID == Student.SID)\
     .filter(Classes.is_deleted == False, Student.is_deleted == False, Marks.assessment_type == 'exam')\
     .group_by(Classes.CName).all()
     
    class_averages = [{'class_name': r[0], 'average': round(float(r[1]), 2)} for r in results if r[1] is not None]
    return render_template("reports/performance.html", class_averages=class_averages)

@reports_bp.route("/reports/student/<int:student_id>/pdf_fast")
@reports_bp.route("/reports/student/<int:student_id>/pdf")
@login_required
def student_report_pdf_fast(student_id):
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'error': 'Student not found'}), 404

    if hasattr(current_user, 'role') and current_user.role == 'teacher':
        from services.teacher_dashboard_service import get_teacher_by_user_id, get_teacher_subject_and_class_ids
        teacher = get_teacher_by_user_id(current_user.id)
        _, class_ids, _ = get_teacher_subject_and_class_ids(teacher)
        if not student.CID or student.CID not in class_ids:
            return jsonify({'error': 'Out-of-scope student pdf report access forbidden'}), 403

    term_id = request.args.get('term_id', type=int)
    exam_id = request.args.get('exam_id', type=int)
    
    query = Marks.query.filter_by(SID=student_id, assessment_type='exam')
    if term_id:
        query = query.filter_by(T_ID=term_id)
    if exam_id:
        query = query.filter_by(ExamID=exam_id)
        
    marks = query.all()
    
    report_data = {}
    for mark in marks:
        exam_name = "امتحان الفصل الدراسي"
        if hasattr(mark, 'exam') and mark.exam:
            exam_name = mark.exam.ExamName
        elif mark.ExamID:
            exam_obj = TypeExams.query.get(mark.ExamID)
            if exam_obj:
                exam_name = exam_obj.ExamName
            
        if exam_name not in report_data:
            report_data[exam_name] = []
        report_data[exam_name].append(mark)
        
    pdf_bytes = generate_student_pdf(student, report_data)
    
    response = make_response(bytes(pdf_bytes))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=report_{student.SID}.pdf'
    return response

@reports_bp.route("/reports/student/<int:student_id>/excel")
@login_required
def student_report_excel(student_id):
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'error': 'Student not found'}), 404

    term_id = request.args.get('term_id', type=int)
    exam_id = request.args.get('exam_id', type=int)
    
    query = Marks.query.filter_by(SID=student_id)
    if term_id:
        query = query.filter_by(T_ID=term_id)
    if exam_id:
        query = query.filter_by(ExamID=exam_id)
        
    marks = query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف درجات"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    ws.append(["رقم المادة", "المادة الدراسية", "الدرجة", "التقدير"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for mark in marks:
        sub_name = mark.subject.SubName if mark.subject else 'غير محدد'
        row = [mark.SubID, sub_name, mark.Score, mark.Grade or '—']
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
            
    ws.column_dimensions['B'].width = 30
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name=f'report_{student.SID}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@reports_bp.route("/reports/export/excel")
@login_required
def export_reports_master_excel():
    class_id = request.args.get('class_id', type=int)
    section_id = request.args.get('section_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    term_id = request.args.get('term_id', type=int)
    student_id = request.args.get('student_id', type=int)
    report_type = request.args.get('report_type', 'class_grades')
    
    query = Student.query.filter_by(is_deleted=False)
    if class_id:
        query = query.filter_by(CID=class_id)
    if section_id:
        query = query.filter_by(SectionID=section_id)
    if student_id:
        query = query.filter_by(SID=student_id)
    
    students = query.order_by(Student.SName).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف التقارير الأكاديمية"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    if report_type == 'attendance_report':
        headers = ["#", "الرقم الأكاديمي", "اسم الطالب", "الصف الدراسي", "الشعبة", "إجمالي الأيام", "أيام الحضور", "أيام الغياب", "نسبة الحضور"]
    elif report_type == 'homework_report':
        headers = ["#", "الرقم الأكاديمي", "اسم الطالب", "الصف الدراسي", "الشعبة", "إجمالي الواجبات", "حالة الإنجاز", "نسبة الإنجاز"]
    else:
        headers = ["#", "الرقم الأكاديمي", "اسم الطالب", "الصف الدراسي", "الشعبة", "عدد المواد", "متوسط الدرجات", "التقدير العام"]
        
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for i, st in enumerate(students, start=1):
        cname = st.school_class.CName if st.school_class else "غير محدد"
        sname = st.section.SectionName if st.section else "غير محدد"
        
        if report_type == 'attendance_report':
            att_q = Attendance.query.filter_by(SID=st.SID)
            tot = att_q.count()
            pres = att_q.filter_by(Status='حاضر').count()
            absent = att_q.filter_by(Status='غائب').count()
            rate = f"{round((pres / tot * 100), 1)}%" if tot > 0 else "100%"
            row = [i, st.SID, st.SName, cname, sname, tot, pres, absent, rate]
        elif report_type == 'homework_report':
            hw_q = Homework.query
            if class_id: hw_q = hw_q.filter_by(class_id=class_id)
            if section_id: hw_q = hw_q.filter_by(section_id=section_id)
            if subject_id: hw_q = hw_q.filter_by(sub_id=subject_id)
            hws = hw_q.all()
            comp = len([h for h in hws if h.status == 'مكتمل'])
            rate = f"{round((comp / len(hws) * 100), 1)}%" if hws else "100%"
            row = [i, st.SID, st.SName, cname, sname, len(hws), f"مكتمل ({comp}/{len(hws)})", rate]
        else:
            st_marks = Marks.query.filter_by(SID=st.SID).all()
            if subject_id:
                st_marks = [m for m in st_marks if m.SubID == subject_id]
            if term_id:
                st_marks = [m for m in st_marks if getattr(m, 'T_ID', None) == term_id]
                
            scores = [float(m.Score) for m in st_marks if m.Score is not None]
            avg = round(sum(scores) / len(scores), 1) if scores else 0.0
            
            if avg >= 90: grade_str = 'ممتاز'
            elif avg >= 80: grade_str = 'جيد جداً'
            elif avg >= 70: grade_str = 'جيد'
            elif avg >= 60: grade_str = 'مقبول'
            else: grade_str = 'دون المستوى' if scores else 'غير مرصود'
            
            row = [i, st.SID, st.SName, cname, sname, len(st_marks), f"{avg}%" if scores else "—", grade_str]
            
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = align_center
            
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'academic_report_{report_type}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route("/reports/export/pdf")
@login_required
def export_reports_master_pdf():
    class_id = request.args.get('class_id', type=int)
    student_id = request.args.get('student_id', type=int)
    
    if student_id:
        student = Student.query.get(student_id)
    elif class_id:
        student = Student.query.filter_by(CID=class_id, is_deleted=False).first()
    else:
        student = Student.query.filter_by(is_deleted=False).first()
        
    if not student:
        return jsonify({'error': 'No student records available for export'}), 404
        
    return student_report_pdf_fast(student.SID)

@reports_bp.route("/reports/api/filter")
def api_reports_filter():
    if not current_user.is_authenticated and not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Unauthorized', 'students': []}), 401
        
    class_id = request.args.get('class_id', type=int)
    section_id = request.args.get('section_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    term_id = request.args.get('term_id', type=int)
    academic_year = request.args.get('academic_year')
    student_id = request.args.get('student_id', type=int)
    report_type = request.args.get('report_type', 'class_grades')
    
    query = Student.query.filter_by(is_deleted=False)
    if class_id:
        query = query.filter_by(CID=class_id)
    if section_id:
        query = query.filter_by(SectionID=section_id)
    if student_id:
        query = query.filter_by(SID=student_id)
        
    students = query.order_by(Student.SName).all()
    
    data = []
    page_target_url = '/reports/performance'
    
    if report_type == 'attendance_report':
        page_target_url = '/attendance'
        for st in students:
            att_q = Attendance.query.filter_by(SID=st.SID)
            tot = att_q.count()
            pres = att_q.filter_by(Status='حاضر').count()
            absent = att_q.filter_by(Status='غائب').count()
            rate = round((pres / tot * 100), 1) if tot > 0 else 100.0
            data.append({
                'student_id': st.SID,
                'name': st.SName,
                'class_name': st.school_class.CName if st.school_class else '—',
                'section_name': st.section.SectionName if st.section else '—',
                'col4': f"{tot} يوم سجل",
                'col5': f"حاضر {pres} | غائب {absent}",
                'average': rate,
                'status_str': f"حضور {rate}% ({pres}/{tot})",
                'badge': 'انضباط ممتاز' if rate >= 90 else ('حضور متوسط' if rate >= 75 else 'غياب مرتفع'),
                'page_url': f"/attendance?class_id={st.CID or ''}&section_id={st.SectionID or ''}"
            })
    elif report_type == 'homework_report':
        page_target_url = '/homework'
        for st in students:
            hw_q = Homework.query
            if class_id: hw_q = hw_q.filter_by(class_id=class_id)
            if section_id: hw_q = hw_q.filter_by(section_id=section_id)
            if subject_id: hw_q = hw_q.filter_by(sub_id=subject_id)
            hws = hw_q.all()
            comp = len([h for h in hws if h.status == 'مكتمل'])
            rate = round((comp / len(hws) * 100), 1) if hws else 100.0
            data.append({
                'student_id': st.SID,
                'name': st.SName,
                'class_name': st.school_class.CName if st.school_class else '—',
                'section_name': st.section.SectionName if st.section else '—',
                'col4': f"{len(hws)} واجب مسجل",
                'col5': f"مكتمل ({comp}/{len(hws)})",
                'average': rate,
                'status_str': f"مكتمل ({comp}/{len(hws)})",
                'badge': 'منجز بالكامل' if rate >= 90 else ('مكتمل جزئياً' if rate >= 60 else 'متأخر'),
                'page_url': '/homework'
            })
    elif report_type == 'exam_report':
        page_target_url = '/exams'
        subjects_list = Subject.query.filter_by(is_deleted=False)
        if subject_id:
            subjects_list = subjects_list.filter_by(SubID=subject_id)
        for sub in subjects_list.all():
            m_list = Marks.query.filter_by(SubID=sub.SubID).all()
            if class_id:
                m_list = [m for m in m_list if m.student and m.student.CID == class_id]
            scores = [float(m.Score) for m in m_list if m.Score is not None]
            avg = round(sum(scores) / len(scores), 1) if scores else 0.0
            max_s = max(scores) if scores else 0
            data.append({
                'student_id': sub.SubID,
                'name': f"اختبارات مادة {sub.SubName}",
                'class_name': 'جميع الصفوف' if not class_id else (Classes.query.get(class_id).CName if Classes.query.get(class_id) else '—'),
                'section_name': 'التقييم الأكاديمي',
                'col4': f"{len(scores)} طالب مرصود",
                'col5': f"أعلى درجة: {max_s} %",
                'average': avg,
                'status_str': f"المعدل العام: {avg}%",
                'badge': 'أداء ممتاز' if avg >= 80 else 'أداء متوسط',
                'page_url': '/exams'
            })
    elif report_type == 'subject_report':
        page_target_url = '/academic/subjects'
        subjects_list = Subject.query.filter_by(is_deleted=False)
        if subject_id:
            subjects_list = subjects_list.filter_by(SubID=subject_id)
        for sub in subjects_list.all():
            m_list = Marks.query.filter_by(SubID=sub.SubID).all()
            scores = [float(m.Score) for m in m_list if m.Score is not None]
            avg = round(sum(scores) / len(scores), 1) if scores else 0.0
            data.append({
                'student_id': sub.SubID,
                'name': sub.SubName,
                'class_name': 'منهج مقر',
                'section_name': 'مادة دراسية',
                'col4': f"{len(m_list)} درجة مرصودة",
                'col5': f"متوسط المادة: {avg}%",
                'average': avg,
                'status_str': f"متوسط المادة: {avg}%",
                'badge': 'مادة أساسية',
                'page_url': '/academic/subjects'
            })
    elif report_type == 'student_grades' and student_id:
        page_target_url = f"/reports/student?student_id={student_id}"
        st = Student.query.get(student_id)
        if st:
            marks = Marks.query.filter_by(SID=st.SID).all()
            if subject_id: marks = [m for m in marks if m.SubID == subject_id]
            if term_id: marks = [m for m in marks if getattr(m, 'T_ID', None) == term_id]
            for m in marks:
                sub_n = m.subject.SubName if m.subject else 'مادة'
                sc = float(m.Score) if m.Score is not None else 0.0
                data.append({
                    'student_id': st.SID,
                    'name': f"{st.SName} — مادة ({sub_n})",
                    'class_name': st.school_class.CName if st.school_class else '—',
                    'section_name': st.section.SectionName if st.section else '—',
                    'col4': f"العظمى: {float(m.MaxScore or 100)}",
                    'col5': f"الدرجة: {sc}",
                    'average': sc,
                    'status_str': f"الدرجة: {sc}",
                    'badge': m.Grade or ('ناجح' if sc >= 50 else 'راسب'),
                    'page_url': f"/reports/student?student_id={st.SID}"
                })
    else:
        if report_type == 'student_grades': page_target_url = '/reports/student'
        elif report_type == 'academic_performance': page_target_url = '/reports/analytics'
        elif report_type == 'class_grades': page_target_url = '/reports/performance'
        
        for st in students:
            marks = Marks.query.filter_by(SID=st.SID).all()
            if subject_id:
                marks = [m for m in marks if m.SubID == subject_id]
            if term_id:
                marks = [m for m in marks if getattr(m, 'T_ID', None) == term_id]
                
            scores = [float(m.Score) for m in marks if m.Score is not None]
            avg = round(sum(scores) / len(scores), 1) if scores else 0.0
            
            if avg >= 90: grade_str = 'ممتاز'
            elif avg >= 80: grade_str = 'جيد جداً'
            elif avg >= 70: grade_str = 'جيد'
            elif avg >= 60: grade_str = 'مقبول'
            else: grade_str = 'متعثر' if scores else 'غير مرصود'

            data.append({
                'student_id': st.SID,
                'name': st.SName,
                'class_name': st.school_class.CName if st.school_class else '—',
                'section_name': st.section.SectionName if st.section else '—',
                'col4': f"{len(marks)} مواد",
                'col5': f"المعدل: {avg}%",
                'average': avg,
                'status_str': f"المعدل: {avg}%",
                'badge': grade_str,
                'page_url': f"/reports/student?student_id={st.SID}&class_id={st.CID or ''}"
            })
            
        if report_type == 'top_students':
            data = [s for s in data if s['average'] >= 80]
            data.sort(key=lambda x: x['average'], reverse=True)
            for idx, s in enumerate(data, start=1):
                s['badge'] = f"المركز #{idx} - متفوق"
        elif report_type == 'struggling_students':
            data = [s for s in data if s['average'] < 60]
            data.sort(key=lambda x: x['average'])
            for s in data:
                s['badge'] = 'يحتاج دعم أكاديمي'
            
    return jsonify({
        'success': True,
        'report_type': report_type,
        'page_target_url': page_target_url,
        'total': len(data),
        'students': data
    })

